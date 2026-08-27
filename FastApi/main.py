# ============================================================
# Central Alert Management System (CAMS) — FastAPI Application
# Production-ready with real-time WebSocket syncing
# ============================================================
import uuid
import os
import time
import json
import hashlib
import asyncio
import logging
import calendar
import io
from contextlib import asynccontextmanager
from datetime import datetime, date, timedelta, timezone
from typing import List, Optional, Dict, Union
from decimal import Decimal

import pytz
import pandas as pd
from fastapi import (
    FastAPI, WebSocket, WebSocketDisconnect, Query, HTTPException,
    Depends, APIRouter,Request
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.encoders import jsonable_encoder
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from jose import jwt, JWTError
from redis import asyncio as redis
from databases import Database
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

from database import database
from database2 import database2
import firebase_admin
from firebase_admin import credentials, messaging

# ============================================================
# Environment Configuration
# ============================================================
load_dotenv()

SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    raise RuntimeError("SECRET_KEY is missing in environment variables.")

ALGORITHM = "HS256"

ALLOWED_ORIGINS = os.getenv(
    "ALLOWED_ORIGINS",
    "http://localhost:5173,http://localhost:3000,http://127.0.0.1:5173,http://127.0.0.1:3000,https://cams.jaesmp.com,http://cams.jaesmp.com"
).split(",")


# ============================================================
# Logging Configuration
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger("central_alerts")


# ============================================================
# Timezone
# ============================================================
ist = pytz.timezone("Asia/Kolkata")
now = datetime.now(ist)
formatted_time = now.strftime("%Y-%m-%d %H:%M:%S")
logger.info(f"Server time: {formatted_time}")


# ============================================================
# Constants
# ============================================================
MAX_CACHE_ENTRIES = 2000
WS_QUEUE_MAX_SIZE = 500


# ============================================================
# Global State
# ============================================================
_cache = {}
_cache_expiry = {}
redis_client = None
connected_clients: dict[str, List[WebSocket]] = {}

alert_worker_task = None
notifier_task = None
redis_sub_task = None
denial_worker_task = None
esc_bump_task = None

# last_sent_updated = None
# last_sent_alert_id = None

security = HTTPBearer()


# ============================================================
# Redis Initialization & Pub/Sub
# ============================================================
async def init_redis():
    """Initialize Redis client (singleton)."""
    global redis_client
    if redis_client is None:
        redis_client = redis.from_url(
            "redis://localhost",
            encoding="utf-8",
            decode_responses=True,
        )
    return redis_client


async def publish_to_redis(channel: str, payload: dict):
    """Publish a message to a Redis channel for cross-worker broadcast."""
    try:
        r = await init_redis()
        await r.publish(channel, json.dumps(payload, default=str))
    except Exception as e:
        logger.error(f"Redis publish failed on channel '{channel}': {e}")


async def redis_subscriber():
    """
    Subscribe to Redis channel and push received messages to local client queues.
    Runs in every worker process to enable cross-worker WebSocket broadcasts.
    """
    logger.info("Redis Subscriber STARTED")
    try:
        r = await init_redis()
        pubsub = r.pubsub()
        await pubsub.subscribe("central_alerts_channel")

        async for message in pubsub.listen():
            if message["type"] == "message":
                try:
                    payload = json.loads(message["data"])

                    # 👇 Direct push to local queues (NOT via manager.broadcast — avoids loop)
                    dead = []
                    for ws, info in manager.active_connections.items():
                        try:
                            info["queue"].put_nowait(payload)
                        except asyncio.QueueFull:
                            logger.warning("Queue full, dropping client")
                            dead.append(ws)
                    for ws in dead:
                        manager.disconnect(ws)

                except Exception as e:
                    logger.error(f"Redis subscriber parse error: {e}")
    except asyncio.CancelledError:
        logger.info("Redis subscriber cancelled")
    except Exception as e:
        logger.exception(f"Redis subscriber error: {e}")


# ============================================================
# Cached Query (Redis + Memory with eviction)
# ============================================================
async def cached_query(sql, params=None, ttl=15, fetch="all", db=database):
    """
    Hybrid cache: local memory + Redis.
    Falls back to direct DB query if both caches miss.
    Automatically evicts oldest entries when cache exceeds MAX_CACHE_ENTRIES.
    """
    await init_redis()

    key_data = {"db": id(db), "sql": sql, "params": params, "fetch": fetch}
    raw_key = json.dumps(key_data, sort_keys=True, default=str)
    cache_key = "cache_query:" + hashlib.md5(raw_key.encode()).hexdigest()

    now = time.time()

    # 1. Local memory cache
    if cache_key in _cache and now < _cache_expiry.get(cache_key, 0):
        return _cache[cache_key]

    # 2. Redis cache
    try:
        redis_data = await redis_client.get(cache_key)
        if redis_data:
            result = json.loads(redis_data)
            _cache[cache_key] = result
            _cache_expiry[cache_key] = now + ttl
            return result
    except Exception as e:
        logger.warning(f"Redis cache read error, falling back to DB: {e}")

    # 3. DB query
    if fetch == "one":
        row = await db.fetch_one(sql, params)
        result = dict(row) if row else None
    else:
        rows = await db.fetch_all(sql, params)
        result = [dict(r) for r in rows]

    # 4. Save to local cache
    _cache[cache_key] = result
    _cache_expiry[cache_key] = now + ttl

    # Evict oldest entries if cache exceeds limit
    if len(_cache) > MAX_CACHE_ENTRIES:
        sorted_keys = sorted(_cache_expiry.items(), key=lambda x: x[1])
        keys_to_remove = [k for k, _ in sorted_keys[:MAX_CACHE_ENTRIES // 4]]
        for k in keys_to_remove:
            _cache.pop(k, None)
            _cache_expiry.pop(k, None)
        logger.info(f"Cache evicted {len(keys_to_remove)} entries (remaining: {len(_cache)})")

    # Save to Redis (best effort)
    try:
        await redis_client.set(cache_key, json.dumps(result, default=str), ex=ttl)
    except Exception:
        pass

    return result


# ============================================================
# JWT Authentication
# ============================================================
def generate_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(hours=24)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def verify_token(creds: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(creds.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        return payload["sub"]
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")


async def verify_jwt_token(token: str):
    if not token:
        return None
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")

        row = await database.fetch_one(
            "SELECT clg_is_login FROM ems_colleague WHERE clg_ref_id = :u",
            {"u": user_id}
        )

        if not row or row["clg_is_login"] != "yes":
            return None

        return user_id
    except JWTError:
        return None


# ============================================================
# Helper Functions
# ============================================================
def format_seconds_to_mmss(total_seconds):
    minutes = int(total_seconds // 60)
    seconds = int(total_seconds % 60)
    return f"{minutes:02}:{seconds:02}"


def format_seconds_to_hhmmss(total_seconds):
    t = timedelta(seconds=int(total_seconds))
    return str(t)


def normalize_row(row):
    """Convert a DB row into a plain dictionary with datetime handling."""
    if not row:
        return {}
    try:
        data = dict(row._mapping)
    except AttributeError:
        data = row if isinstance(row, dict) else {}

    normalized = {}
    for k, v in data.items():
        if v is None:
            normalized[k] = None
        elif isinstance(v, (datetime, date)):
            normalized[k] = v.strftime("%Y-%m-%d %H:%M:%S")
        else:
            normalized[k] = str(v)
    return normalized


def serialize_row(row):
    """Convert a database row to a JSON-serializable dict."""
    if hasattr(row, '_mapping'):
        data = dict(row._mapping)
    elif hasattr(row, 'keys'):
        data = {key: row[key] for key in row.keys()}
    else:
        data = dict(row)

    for k, v in list(data.items()):
        if isinstance(v, (datetime, date)):
            data[k] = v.isoformat()
        elif isinstance(v, Decimal):
            data[k] = float(v)
    return data


def to_float(val):
    try:
        return float(val) if val not in ("", None) else None
    except Exception:
        return None


def to_datetime(val):
    try:
        if not val:
            return None
        if isinstance(val, datetime):
            return val
        return datetime.strptime(val, "%Y-%m-%d %H:%M:%S")
    except Exception:
        return None


def to_int(val):
    try:
        return int(val) if val not in ("", None) else None
    except Exception:
        return None


def hhmmss_to_seconds(value: str) -> int:
    """Convert HH:MM:SS text to seconds safely. Returns 0 for invalid values."""
    try:
        if not value:
            return 0
        h, m, s = value.split(":")
        return int(h) * 3600 + int(m) * 60 + int(s)
    except Exception:
        return 0

def format_seconds_human(seconds):
    """Seconds ko human-readable format me convert karta hai.
    
    Examples:
        45     → "45 sec"
        90     → "1 min 30 sec"
        300    → "5 min"
        3600   → "1 hr"
        3900   → "1 hr 5 min"
        2880   → "48 min"
    """
    if seconds is None:
        return "N/A"

    seconds = int(seconds)

    if seconds < 60:
        return f"{seconds} sec"

    minutes = seconds // 60
    remaining_seconds = seconds % 60

    if minutes < 60:
        parts = [f"{minutes} min"]
        if remaining_seconds > 0:
            parts.append(f"{remaining_seconds} sec")
        return " ".join(parts)

    hours = minutes // 60
    remaining_minutes = minutes % 60

    parts = [f"{hours} hr"]
    if remaining_minutes > 0:
        parts.append(f"{remaining_minutes} min")
    return " ".join(parts)


def hash_filter(data: dict):
    return json.dumps(data, sort_keys=True)


def group_by_severity(rows):
    """Group alerts by severity."""
    severity_order = ["CRITICAL", "HIGH", "MEDIUM", "LOW"]
    grouped = {sev: [] for sev in severity_order}

    for r in rows:
        serialized = serialize_row(r)
        sev = (serialized.get("severity") or "").upper()
        if sev in grouped:
            grouped[sev].append(serialized)
        else:
            grouped["LOW"].append(serialized)
    return grouped


def get_date_filter(range_type: str):
    if range_type == "today":
        return "DATE(created_date) = CURRENT_DATE"
    elif range_type == "month":
        return "DATE_TRUNC('month', created_date) = DATE_TRUNC('month', CURRENT_DATE)"
    else:
        return "1=1"


def format_worksheet(ws):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 3


class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        elif isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)


# ============================================================
# Pydantic Models
# ============================================================
class LoginRequest(BaseModel):
    username: str
    password: str


class LogoutRequest(BaseModel):
    username: str


class DistrictOut(BaseModel):
    district_id: int
    district_name: str


class DivisionOut(BaseModel):
    division_id: int
    division_name: str


class EscalateRequest(BaseModel):
    remark: Optional[str] = None
    escalated_by: Optional[str] = None


class SeverityUpdate(BaseModel):
    alert_id: int
    severity: str


class CancelUpdate(BaseModel):
    alert_id: int
    remark: str
    cancel_by: Optional[str] = None


class AlertThresholdUpdate(BaseModel):
    threshold_seconds: Optional[int] = None
    severity: Optional[str] = None
    priority: Optional[int] = None


# ============================================================
# Alert Thresholds & Worker
# ============================================================
async def get_active_thresholds():
    """Fetch active thresholds, cached for 30s."""
    query = """
        SELECT alert_type, amb_area, threshold_seconds, severity, priority
        FROM alert_thresholds
        WHERE is_active = TRUE
        ORDER BY priority ASC;
    """
    rows = await cached_query(query, fetch="all", ttl=30, db=database2)
    return [normalize_row(r) for r in rows]


def resolve_alerts(row, thresholds):
    acknowledge_sec = hhmmss_to_seconds(row.get("acknowledge_duration"))
    start_base_sec = hhmmss_to_seconds(row.get("start_from_base_duration"))
    at_scene_sec = hhmmss_to_seconds(row.get("at_scene_duration"))
    amb_area = row.get("amb_working_area")

    ack_raw = row.get("acknowledge_duration")
    ack_done = ack_raw is not None and str(ack_raw).strip() != ""

    patient_handover_dt = to_datetime(row.get("patient_handover"))
    back_to_base_dt = to_datetime(row.get("back_to_base_loc"))

    back_to_base_sec = None
    if patient_handover_dt and back_to_base_dt:
        diff = (back_to_base_dt - patient_handover_dt).total_seconds()
        if diff >= 0:
            back_to_base_sec = diff

    pilot_login_out_val = row.get("pilot_login_out")
    mdt_not_found = pilot_login_out_val is None or pilot_login_out_val == "No"

    metric_map = {
        "ACK_DELAY": acknowledge_sec,
        "START_DELAY": start_base_sec,
        "AT_SCENE_DELAY": at_scene_sec,
        "BACK_TO_BASE_DELAY": back_to_base_sec,
    }

    matched_alerts = []
    seen_alert_types = set()

    for t in thresholds:
        if t["amb_area"] is not None and t["amb_area"] != amb_area:
            continue

        alert_type = t["alert_type"]

        if alert_type in seen_alert_types:
            continue

        if alert_type == "MDT_NOT_LOGGED_IN":
            if ack_done and mdt_not_found:
                matched_alerts.append((alert_type, t["severity"]))
                seen_alert_types.add(alert_type)
            continue

        metric_value = metric_map.get(alert_type)
        threshold_val = int(t["threshold_seconds"])

        if metric_value is not None and metric_value > threshold_val:
            matched_alerts.append((alert_type, t["severity"]))
            seen_alert_types.add(alert_type)

    return matched_alerts


async def rtm_alert_insert_worker():
    """Background worker that checks RTM dashboard and inserts alerts."""
    logger.info("RTM Alert Insert Worker STARTED")

    while True:
        try:
            query = """
                SELECT *
                FROM rtm_dashboard
                WHERE inc_datetime >= CURRENT_DATE
                  AND inc_datetime < CURRENT_DATE + INTERVAL '1 day'
                ORDER BY inc_datetime DESC
                LIMIT 1000;
            """

            rows = await cached_query(query, fetch="all", ttl=5, db=database2)
            thresholds = await get_active_thresholds()

            for row in rows:
                try:
                    row = normalize_row(row)
                    alerts = resolve_alerts(row, thresholds)
                    if not alerts:
                        continue

                    for alert_type, severity in alerts:
                        params = {
                            "alert_type": alert_type,
                            "incident_id": row.get("inc_ref_id"),
                            "system_type": row.get("inc_system_type"),
                            "severity": severity,
                            "ambulance_no": row.get("ambulance_no"),
                            "remark": f"{alert_type} threshold breached",
                            "division": row.get("division_name"),
                            "district": row.get("district_name"),
                            "inc_latitude": to_float(row.get("inc_lat")),
                            "inc_longitude": to_float(row.get("inc_long")),
                            "amb_lat": to_float(row.get("gps_amb_lat")),
                            "amb_long": to_float(row.get("gps_amb_log")),
                            "inc_datetime": to_datetime(row.get("inc_datetime")),
                            "pilot_name": row.get("pilot_name"),
                            "pilot_mobile": to_int(row.get("pilot_mobile")),
                            "paramedic_name": row.get("paramedic_name"),
                            "paramedic_mobile": to_int(row.get("paramedic_mobile")),
                        }

                        # 👇 NAYA: fetch_one + RETURNING use karo taaki pata chale insert hua ya nahi
                        inserted_row = await database2.fetch_one(
                            """
                            INSERT INTO central_alerts (
                                alert_type, incident_id, system_type, severity,
                                ambulance_no, remark, division, district,
                                inc_latitude, inc_longitude, amb_lat, amb_long,
                                inc_datetime, pilot_name, pilot_mobile,
                                paramedic_name, paramedic_mobile
                            )
                            VALUES (
                                :alert_type, :incident_id, :system_type, :severity,
                                :ambulance_no, :remark, :division, :district,
                                :inc_latitude, :inc_longitude, :amb_lat, :amb_long,
                                :inc_datetime, :pilot_name, :pilot_mobile,
                                :paramedic_name, :paramedic_mobile
                            )
                            ON CONFLICT (incident_id, alert_type) DO NOTHING
                            RETURNING *
                            """,
                            params
                        )

                        # 👇 Agar insert hua (conflict nahi hua) → broadcast karo
                        if inserted_row:
                            inserted_dict = serialize_row(inserted_row)
                            try:
                                await broadcast_new_central_alert(inserted_dict)
                            except Exception as b_err:
                                logger.error(
                                    f"Broadcast failed for incident_id={params.get('incident_id')}: {b_err}"
                                )

                except Exception as row_err:
                    logger.warning(f"Row error (incident_id={row.get('inc_ref_id')}): {row_err}")
                    continue

        except Exception as e:
            logger.error(f"RTM Alert Worker Error: {e}")

        await asyncio.sleep(5)
    ######################################################################################

# import uuid

# # Columns jo hum MySQL se map kar rahe hain
# MAPPED_COLUMNS = {
#     "amb_no", "amb_default_mobile", "caller_no",
#     "hp_name", "challenge_val", "meaning", "denial_remark",
#     "added_by", "added_date", "call_id"        # 👈 call_id add kiya
# }

# _required_defaults_cache = None

# DENIAL_WORKER_LOCK_KEY = "denial_worker_leader_lock"
# DENIAL_WORKER_LOCK_TTL_MS = 15000
# _worker_instance_id = str(uuid.uuid4())


# async def try_acquire_denial_worker_leadership() -> bool:
#     """Redis leader lock — sirf ek process actual kaam kare."""
#     acquired = await redis_client.set(
#         DENIAL_WORKER_LOCK_KEY,
#         _worker_instance_id,
#         nx=True,
#         px=DENIAL_WORKER_LOCK_TTL_MS
#     )
#     if acquired:
#         return True

#     current_holder = await redis_client.get(DENIAL_WORKER_LOCK_KEY)
#     if isinstance(current_holder, bytes):
#         current_holder = current_holder.decode()

#     if current_holder == _worker_instance_id:
#         await redis_client.pexpire(DENIAL_WORKER_LOCK_KEY, DENIAL_WORKER_LOCK_TTL_MS)
#         return True

#     return False


# async def get_required_default_columns():
#     """Postgres ke NOT NULL (bina default) columns ke liye safe defaults."""
#     global _required_defaults_cache

#     if _required_defaults_cache is not None:
#         return _required_defaults_cache

#     rows = await database2.fetch_all(
#         """
#         SELECT column_name, data_type
#         FROM information_schema.columns
#         WHERE table_schema = 'public'
#           AND table_name = 'denial_escalation_master'
#           AND is_nullable = 'NO'
#           AND column_default IS NULL
#         """
#     )

#     defaults = {}
#     for r in rows:
#         col = r["column_name"]
#         dtype = r["data_type"]

#         # in columns ko hum khud fill karte hain
#         if col in MAPPED_COLUMNS or col in ("id", "mysql_id"):
#             continue

#         if dtype in ("integer", "bigint", "smallint", "numeric"):
#             defaults[col] = 0
#         elif dtype == "boolean":
#             defaults[col] = False
#         elif dtype in ("character varying", "text", "varchar", "char"):
#             defaults[col] = ""
#         elif dtype in ("timestamp without time zone", "timestamp with time zone", "date"):
#             defaults[col] = datetime.utcnow()
#         else:
#             defaults[col] = None

#     _required_defaults_cache = defaults
#     logger.info(f"Denial worker: required NOT NULL defaults resolved -> {defaults}")
#     return defaults


# async def denial_complaints_insert_worker():
#     """
#     Background worker: MySQL (ems_denial_complaints) -> Postgres (denial_escalation_master).
#     Current month only. Uses MySQL `id` for uniqueness check.
#     inc_ref_id is intentionally SKIPPED. call_id is included.
#     """
#     logger.info(f"Denial Complaints Insert Worker STARTED (instance={_worker_instance_id})")

#     while True:
#         try:
#             is_leader = await try_acquire_denial_worker_leadership()

#             if not is_leader:
#                 await asyncio.sleep(5)
#                 continue

#             required_defaults = await get_required_default_columns()

#             # ------------------------------------------------
#             # STEP 1: MySQL se current month rows lao
#             # ------------------------------------------------
#             query = """
#                 SELECT
#                     id, call_id, amb_no, amb_default_mobile, caller_no,
#                     hp_name, challenge_val, meaning, denial_remark,
#                     added_by, added_date
#                 FROM ems_denial_complaints
#                 WHERE YEAR(added_date) = YEAR(NOW())
#                   AND MONTH(added_date) = MONTH(NOW())
#                 ORDER BY added_date DESC
#                 LIMIT 1000;
#             """

#             rows = await database.fetch_all(query)

#             logger.info("=" * 60)
#             logger.info(f"Denial worker: MySQL returned {len(rows)} rows (current month filter)")

#             if len(rows) == 0:
#                 try:
#                     sample = await database.fetch_one(
#                         "SELECT MIN(added_date) AS min_dt, MAX(added_date) AS max_dt, "
#                         "COUNT(*) AS total FROM ems_denial_complaints"
#                     )
#                     if sample:
#                         logger.warning(
#                             f"Denial worker: 0 rows for current month! "
#                             f"Table has min_date={sample['min_dt']}, "
#                             f"max_date={sample['max_dt']}, total_rows={sample['total']}"
#                         )
#                 except Exception as e:
#                     logger.error(f"Denial worker: sample date check failed: {e}")
#                 await asyncio.sleep(10)
#                 continue

#             # ------------------------------------------------
#             # STEP 2: Sample 3 rows ka data log karo
#             # ------------------------------------------------
#             for idx, r in enumerate(rows[:3]):
#                 try:
#                     d = dict(r)
#                     logger.info(
#                         f"Denial worker: sample row {idx} -> "
#                         f"id={d.get('id')!r}, "
#                         f"call_id={d.get('call_id')!r}, "
#                         f"amb_no={d.get('amb_no')!r}, "
#                         f"challenge_val={d.get('challenge_val')!r}, "
#                         f"added_date={d.get('added_date')!r}"
#                     )
#                 except Exception as e:
#                     logger.error(f"Denial worker: sample log error: {e}")

#             # ------------------------------------------------
#             # STEP 3: Per-row processing
#             # ------------------------------------------------
#             inserted_count = 0
#             skipped_existing = 0
#             skipped_no_id = 0
#             failed_insert = 0
#             failed_examples = []

#             for row in rows:
#                 try:
#                     d = dict(row)
#                     mysql_id = d.get("id")

#                     if not mysql_id:
#                         skipped_no_id += 1
#                         continue

#                     # ------------------------------------------------
#                     # Check karo: ye mysql_id pehle se Postgres mein hai?
#                     # ------------------------------------------------
#                     existing = await database2.fetch_one(
#                         """
#                         SELECT id
#                         FROM denial_escalation_master
#                         WHERE mysql_id = :mysql_id
#                         """,
#                         {"mysql_id": mysql_id}
#                     )

#                     if existing:
#                         skipped_existing += 1
#                         continue

#                     # ------------------------------------------------
#                     # Insert karlo
#                     # ------------------------------------------------
#                     params = {
#                         "id": mysql_id,
#                         "mysql_id": mysql_id,
#                         "call_id": d.get("call_id"),                       # 👈 call_id add kiya
#                         "amb_no": d.get("amb_no"),
#                         "amb_default_mobile": d.get("amb_default_mobile"),
#                         "caller_no": d.get("caller_no"),
#                         "hp_name": d.get("hp_name"),
#                         "challenge_val": d.get("challenge_val"),
#                         "meaning": d.get("meaning"),
#                         "denial_remark": d.get("denial_remark"),
#                         "added_by": d.get("added_by"),
#                         "added_date": d.get("added_date"),
#                     }
#                     params.update(required_defaults)

#                     columns = ", ".join(params.keys())
#                     placeholders = ", ".join(f":{k}" for k in params.keys())

#                     await database2.execute(
#                         f"""
#                         INSERT INTO denial_escalation_master ({columns})
#                         VALUES ({placeholders})
#                         """,
#                         params
#                     )
#                     inserted_count += 1

#                 except Exception as row_err:
#                     failed_insert += 1
#                     if len(failed_examples) < 5:
#                         try:
#                             d_err = dict(row)
#                             failed_examples.append({
#                                 "id": d_err.get("id"),
#                                 "call_id": d_err.get("call_id"),
#                                 "amb_no": d_err.get("amb_no"),
#                                 "error": str(row_err),
#                             })
#                         except Exception:
#                             failed_examples.append({"error": str(row_err)})
#                     continue

#             # ------------------------------------------------
#             # STEP 4: Final summary
#             # ------------------------------------------------
#             logger.info(
#                 f"Denial worker SUMMARY: "
#                 f"fetched={len(rows)}, "
#                 f"inserted={inserted_count}, "
#                 f"existing={skipped_existing}, "
#                 f"no_id={skipped_no_id}, "
#                 f"failed={failed_insert}"
#             )

#             if failed_examples:
#                 logger.warning(
#                     f"Denial worker: failed insert examples -> {failed_examples}"
#                 )

#             if inserted_count:
#                 logger.info(f"Denial worker: ✅ inserted {inserted_count} new rows")

#         except Exception as e:
#             logger.error(f"Denial Complaints Insert Worker Error: {e}", exc_info=True)

#         await asyncio.sleep(10)



# Columns jo hum MySQL se map kar rahe hain
MAPPED_COLUMNS = {
    "amb_no", "amb_default_mobile", "caller_no",
    "hp_name", "challenge_val", "meaning", "denial_remark",
    "added_by", "added_date", "call_id"
}
 
# Forced defaults — ye columns hamesha isi value se bharenge
FORCED_DEFAULTS = {
    "alert_type": "incident denial",
}
 
_required_defaults_cache = None
 
DENIAL_WORKER_LOCK_KEY = "denial_worker_leader_lock"
DENIAL_WORKER_LOCK_TTL_MS = 15000
_worker_instance_id = str(uuid.uuid4())
 
 
async def try_acquire_denial_worker_leadership() -> bool:
    """Redis leader lock — sirf ek process actual kaam kare."""
    acquired = await redis_client.set(
        DENIAL_WORKER_LOCK_KEY,
        _worker_instance_id,
        nx=True,
        px=DENIAL_WORKER_LOCK_TTL_MS
    )
    if acquired:
        return True
 
    current_holder = await redis_client.get(DENIAL_WORKER_LOCK_KEY)
    if isinstance(current_holder, bytes):
        current_holder = current_holder.decode()
 
    if current_holder == _worker_instance_id:
        await redis_client.pexpire(DENIAL_WORKER_LOCK_KEY, DENIAL_WORKER_LOCK_TTL_MS)
        return True
 
    return False
 
 
async def get_required_default_columns():
    """Postgres ke NOT NULL (bina default) columns ke liye safe defaults."""
    global _required_defaults_cache
 
    if _required_defaults_cache is not None:
        return _required_defaults_cache
 
    rows = await database2.fetch_all(
        """
        SELECT column_name, data_type
        FROM information_schema.columns
        WHERE table_schema = 'public'
          AND table_name = 'denial_escalation_master'
          AND is_nullable = 'NO'
          AND column_default IS NULL
        """
    )
 
    defaults = {}
    for r in rows:
        col = r["column_name"]
        dtype = r["data_type"]
 
        # in columns ko hum khud fill karte hain
        # FORCED_DEFAULTS bhi skip — warna alert_type ko "" mil jata
        if col in MAPPED_COLUMNS or col in ("id", "mysql_id") or col in FORCED_DEFAULTS:
            continue
 
        if dtype in ("integer", "bigint", "smallint", "numeric"):
            defaults[col] = 0
        elif dtype == "boolean":
            defaults[col] = False
        elif dtype in ("character varying", "text", "varchar", "char"):
            defaults[col] = ""
        elif dtype in ("timestamp without time zone", "timestamp with time zone", "date"):
            defaults[col] = datetime.utcnow()
        else:
            defaults[col] = None
 
    _required_defaults_cache = defaults
    logger.info(f"Denial worker: required NOT NULL defaults resolved -> {defaults}")
    return defaults
 
 
async def denial_complaints_insert_worker():
    """
    Background worker: MySQL (ems_denial_complaints) -> Postgres (denial_escalation_master).
    """
    logger.info(f"Denial Complaints Insert Worker STARTED (instance={_worker_instance_id})")
    while True:
        try:
            is_leader = await try_acquire_denial_worker_leadership()
            if not is_leader:
                await asyncio.sleep(5)
                continue
            required_defaults = await get_required_default_columns()
            # ------------------------------------------------
            # STEP 1: MySQL se current month rows lao (WITH JOINS)
            # ------------------------------------------------
            # 👇 NAYA QUERY: dst_id ki jagah dst_code use kiya gaya hai
            query = """
                SELECT
                    edc.id, edc.call_id, edc.amb_no, edc.amb_default_mobile, edc.caller_no,
                    edc.hp_name, edc.challenge_val, edc.denial_remark,
                    edc.added_by, edc.added_date,
                    dr.meaning AS reason_meaning,
                    dist.dst_name AS dst_name
                FROM ems_denial_complaints edc
                LEFT JOIN ems_denial_reason dr 
                    ON edc.meaning = dr.id
                LEFT JOIN ems_mas_districts dist 
                    ON edc.amb_district = dist.dst_code  -- 👈 YAHAN FIX KIYA GAYA HAI
                WHERE YEAR(edc.added_date) = YEAR(NOW())
                  AND MONTH(edc.added_date) = MONTH(NOW())
                ORDER BY edc.added_date DESC
                LIMIT 1000;
            """
            rows = await database.fetch_all(query)


            logger.info("=" * 60)

            logger.info(f"Denial worker: MySQL returned {len(rows)} rows (current month filter)")

            if len(rows) == 0:

                try:

                    sample = await database.fetch_one(

                        "SELECT MIN(added_date) AS min_dt, MAX(added_date) AS max_dt, "

                        "COUNT(*) AS total FROM ems_denial_complaints"

                    )

                    if sample:

                        logger.warning(

                            f"Denial worker: 0 rows for current month! "

                            f"Table has min_date={sample['min_dt']}, "

                            f"max_date={sample['max_date']}, total_rows={sample['total']}"

                        )

                except Exception as e:

                    logger.error(f"Denial worker: sample date check failed: {e}")

                await asyncio.sleep(10)

                continue

            # ------------------------------------------------

            # STEP 2: Sample 3 rows ka data log karo

            # ------------------------------------------------

            for idx, r in enumerate(rows[:3]):

                try:

                    d = dict(r)

                    logger.info(

                        f"Denial worker: sample row {idx} -> "

                        f"id={d.get('id')!r}, "

                        f"call_id={d.get('call_id')!r}, "

                        f"amb_no={d.get('amb_no')!r}, "

                        f"reason_meaning={d.get('reason_meaning')!r}, " # 👈 Changed

                        f"dst_name={d.get('dst_name')!r}, "            # 👈 Changed

                        f"added_date={d.get('added_date')!r}"

                    )

                except Exception as e:

                    logger.error(f"Denial worker: sample log error: {e}")

            # ------------------------------------------------

            # STEP 3: Per-row processing

            # ------------------------------------------------

            inserted_count = 0

            skipped_existing = 0

            skipped_no_id = 0

            failed_insert = 0

            failed_examples = []

            for row in rows:

                try:

                    d = dict(row)

                    mysql_id = d.get("id")

                    if not mysql_id:

                        skipped_no_id += 1

                        continue

                    # Check karo: ye mysql_id pehle se Postgres mein hai?

                    existing = await database2.fetch_one(

                        """

                        SELECT id

                        FROM denial_escalation_master

                        WHERE mysql_id = :mysql_id

                        """,

                        {"mysql_id": mysql_id}

                    )

                    if existing:

                        skipped_existing += 1

                        continue

                    # Insert karlo

                    # 👇 NAYA: meaning ki jagah reason_meaning, aur dst_name add kiya

                    params = {

                        "id": mysql_id,

                        "mysql_id": mysql_id,

                        "call_id": d.get("call_id"),

                        "amb_no": d.get("amb_no"),

                        "amb_default_mobile": d.get("amb_default_mobile"),

                        "caller_no": d.get("caller_no"),

                        "hp_name": d.get("hp_name"),

                        "challenge_val": d.get("challenge_val"),

                        "meaning": d.get("reason_meaning"),  # 👈 Ab ye text aayega ID nahi

                        "dst_name": d.get("dst_name"),       # 👈 District naam add kiya

                        "denial_remark": d.get("denial_remark"),

                        "added_by": d.get("added_by"),

                        "added_date": d.get("added_date"),

                    }

                    # pehle NOT NULL defaults lagao

                    params.update(required_defaults)

                    # ab forced defaults — alert_type = "incident deny" hamesha

                    params.update(FORCED_DEFAULTS)

                    columns = ", ".join(params.keys())

                    placeholders = ", ".join(f":{k}" for k in params.keys())

                    await database2.execute(

                        f"""

                        INSERT INTO denial_escalation_master ({columns})

                        VALUES ({placeholders})

                        """,

                        params

                    )

                    inserted_count += 1

                    # NAYA: alert_escalation_flow table me insert karo

                    try:

                        await insert_into_escalation_flow('denial', d)

                    except Exception as flow_err:

                        logger.error(f"Flow insert failed for denial: {flow_err}")

                    # 👈 Separate try/except — broadcast fail ho to insert count disturb na ho

                    try:

                        await broadcast_new_escalation(

                            call_id=d.get("call_id"),

                            denial_record={

                                "id":                mysql_id,

                                "mysql_id":          mysql_id,

                                "call_id":           d.get("call_id"),

                                "amb_no":            d.get("amb_no"),

                                "amb_default_mobile": d.get("amb_default_mobile"),

                                "caller_no":         d.get("caller_no"),

                                "hp_name":           d.get("hp_name"),

                                "challenge_val":     d.get("challenge_val"),

                                "meaning":           d.get("reason_meaning"), # 👈 Text bheja

                                "dst_name":          d.get("dst_name"),       # 👈 Naam bheja

                                "denial_remark":     d.get("denial_remark"),

                                "alert_type":        "incident denial",

                                "added_by":          d.get("added_by"),

                                "added_date":        d.get("added_date").isoformat() if d.get("added_date") else None,

                            },

                        )

                    except Exception as b_err:

                        logger.error(f"Denial worker: broadcast failed for id={mysql_id}: {b_err}")

                except Exception as row_err:

                    failed_insert += 1

                    if len(failed_examples) < 5:

                        try:

                            d_err = dict(row)

                            failed_examples.append({

                                "id": d_err.get("id"),

                                "call_id": d_err.get("call_id"),

                                "amb_no": d_err.get("amb_no"),

                                "error": str(row_err),

                            })

                        except Exception:

                            failed_examples.append({"error": str(row_err)})

                    continue

            # ------------------------------------------------

            # STEP 4: Final summary

            # ------------------------------------------------

            logger.info(

                f"Denial worker SUMMARY: "

                f"fetched={len(rows)}, "

                f"inserted={inserted_count}, "

                f"existing={skipped_existing}, "

                f"no_id={skipped_no_id}, "

                f"failed={failed_insert}"

            )

            if failed_examples:

                logger.warning(

                    f"Denial worker: failed insert examples -> {failed_examples}"

                )

            if inserted_count:

                logger.info(f"Denial worker: ✅ inserted {inserted_count} new rows")

        except Exception as e:

            logger.error(f"Denial Complaints Insert Worker Error: {e}", exc_info=True)

        await asyncio.sleep(10)
 
    
    ##################################################################################


# ============================================================
# Connection Manager
# ============================================================
class ConnectionManager:
    """
    Each connection gets its own asyncio.Queue.
    Broadcaster pushes to all queues (non-blocking).
    Each WS worker drains its own queue so slow clients don't block others.
    """
    def __init__(self):
        self.active_connections: dict[WebSocket, dict] = {}

    async def connect(self, websocket: WebSocket, user_id: str):
        await websocket.accept()
        self.active_connections[websocket] = {
            "user_id": user_id,
            "queue": asyncio.Queue(maxsize=WS_QUEUE_MAX_SIZE),
            "last_seen": time.time(),
        }
        return self.active_connections[websocket]

    def disconnect(self, websocket: WebSocket):
        self.active_connections.pop(websocket, None)

    async def safe_send(self, websocket: WebSocket, payload: dict):
        try:
            await websocket.send_json(payload)
            return True
        except Exception as e:
            logger.warning(f"Send failed, disconnecting: {e}")
            self.disconnect(websocket)
            return False

    def broadcast(self, payload: dict):
        """Publish to Redis — ALL workers (including this one) will receive
        via redis_subscriber and push to their local client queues.
        This ensures cross-worker WebSocket broadcasts work correctly."""
        try:
            asyncio.create_task(publish_to_redis("central_alerts_channel", payload))
        except Exception as e:
            logger.error(f"Broadcast (Redis publish) failed: {e}")


manager = ConnectionManager()


# ============================================================
# Central Alerts Payload Helper
# ============================================================
async def _fetch_alerts_payload(incident_id=None, filter_date=None):
    """
    Fetch alerts + counts as a single payload.
    Used for initial connect, filter requests, and broadcast updates.
    Always returns data sorted by inc_datetime DESC (order guaranteed).
    """
    conditions = []
    params = {}

    if incident_id:
        conditions.append("incident_id = :incident_id")
        params["incident_id"] = str(incident_id)
    elif filter_date:
        start_dt = datetime.strptime(filter_date, "%Y-%m-%d")
        end_dt = start_dt + timedelta(days=1)
        conditions.append("created_date >= :start_date AND created_date < :end_date")
        params["start_date"] = start_dt
        params["end_date"] = end_dt
    else:
        conditions.append("inc_datetime >= CURRENT_DATE AND inc_datetime < CURRENT_DATE + INTERVAL '1 day'")

    where_clause = " AND ".join(conditions)

    query = f"""
        SELECT * FROM central_alerts
        WHERE {where_clause}
        ORDER BY inc_datetime DESC
    """

    if filter_date:
        rows = await database2.fetch_all(query, params)
    else:
        rows = await cached_query(query, params=params, ttl=3, fetch="all", db=database2)

    today_all = [serialize_row(r) for r in rows]
    by_severity = group_by_severity(rows)

    # Counts recalculated for every filter
    count_rows = await cached_query(
        f"""
        SELECT system_type, severity, COUNT(*) AS total
        FROM central_alerts
        WHERE {where_clause}
          AND escalate_status = '1'
          AND system_type IN ('108', '102')
        GROUP BY system_type, severity
        """,
        params=params,
        ttl=3,
        fetch="all",
        db=database2,
    ) or []

    counts = {"total": {"108": 0, "102": 0}, "severity": {"108": {}, "102": {}}}
    for r in count_rows:
        sys_t = r["system_type"]
        sev = r["severity"]
        counts["total"][sys_t] += r["total"]
        counts["severity"][sys_t][sev] = r["total"]

    return {
        "type": "ALL_ALERTS",
        "data": {
            "today_all": today_all,
            "by_severity": by_severity,
            "counts": counts,
        },
    }


# ============================================================
# Alert WebSocket Notifier (with Redis pub/sub)
# ============================================================
# ===============================
# ALERT CHANGE NOTIFIER
# ===============================

last_sent_updated = None
last_sent_alert_id = None


async def alert_ws_notifier():
    global last_sent_updated, last_sent_alert_id
    print("🚀 Alert WebSocket Notifier STARTED")

    while True:
        try:
            if last_sent_updated is None:
                row = await database2.fetch_one(
                    """
                    SELECT alert_id, COALESCE(updated_date, created_date) AS last_time
                    FROM central_alerts
                    ORDER BY COALESCE(updated_date, created_date) DESC, alert_id DESC
                    LIMIT 1
                    """
                )
                if row:
                    last_sent_updated = row["last_time"]
                    last_sent_alert_id = row["alert_id"]

            else:
                rows = await database2.fetch_all(
                    """
                    SELECT *
                    FROM central_alerts
                    WHERE (COALESCE(updated_date, created_date) > :last_time)
                       OR (COALESCE(updated_date, created_date) = :last_time AND alert_id > :last_id)
                    ORDER BY COALESCE(updated_date, created_date) ASC, alert_id ASC
                    """,
                    {
                        "last_time": last_sent_updated,
                        "last_id": last_sent_alert_id
                    }
                )

                if rows:
                    last_row = rows[-1]
                    last_sent_updated = last_row["updated_date"] if last_row["updated_date"] else last_row["created_date"]
                    last_sent_alert_id = last_row["alert_id"]

                    # 👇 FULL payload bhejo ALL_ALERTS format me (same as initial load)
                    full_payload = await _fetch_alerts_payload()
                    manager.broadcast(full_payload)
                    print(f"📡 WS sent ALL_ALERTS update ({len(rows)} changes detected)")

        except Exception as e:
            print("❌ Alert WS Notifier Error:", e)

        # 👇 0.5 sec — pehle 1 sec tha, ab fast hoga
        await asyncio.sleep(0.5)


# ============================================================
# Auto-Restart Wrappers (crash recovery)
# ============================================================
async def run_notifier_with_restart():
    """Wrapper that restarts alert_ws_notifier if it crashes."""
    while True:
        try:
            logger.info("Starting alert_ws_notifier...")
            await alert_ws_notifier()
        except asyncio.CancelledError:
            logger.info("Notifier cancelled, exiting wrapper")
            break
        except Exception as e:
            logger.exception(f"Notifier CRASHED — restarting in 5s: {e}")
            await asyncio.sleep(5)


async def run_worker_with_restart():
    """Wrapper that restarts rtm_alert_insert_worker if it crashes."""
    while True:
        try:
            logger.info("Starting rtm_alert_insert_worker...")
            await rtm_alert_insert_worker()
        except asyncio.CancelledError:
            logger.info("Worker cancelled, exiting wrapper")
            break
        except Exception as e:
            logger.exception(f"Worker CRASHED — restarting in 5s: {e}")
            await asyncio.sleep(5)


# ============================================================
# Lifespan (Startup & Shutdown)
# ============================================================
@asynccontextmanager
async def lifespan(app: FastAPI):
    global alert_worker_task, notifier_task, redis_sub_task, denial_worker_task
    global esc_bump_task

    # --- STARTUP ---
    await database.connect()
    await database2.connect()
    await init_redis()

    # 👇 Firebase Admin SDK initialize karo — workers start hone se PEHLE
    # (kyunki workers FCM push bhejte hain — Firebase ready hona chahiye)
    try:
        firebase_admin.get_app()  # Already initialized?
        logger.info("Firebase Admin SDK already initialized")
    except ValueError:
        try:
            cred_path = os.getenv("FIREBASE_CREDENTIALS_PATH", "/home/jaesadmin/firebase-creds.json")
            cred = credentials.Certificate(cred_path)
            firebase_admin.initialize_app(cred)
            logger.info(f"Firebase Admin SDK initialized with {cred_path}")
        except Exception as e:
            logger.error(f"Firebase Admin init failed: {e}")

    alert_worker_task = asyncio.create_task(run_worker_with_restart())
    notifier_task = asyncio.create_task(run_notifier_with_restart())
    redis_sub_task = asyncio.create_task(redis_subscriber())
    denial_worker_task = asyncio.create_task(denial_complaints_insert_worker())
    esc_bump_task = asyncio.create_task(escalation_level_bump_watcher())

    logger.info("Application STARTED — all workers + escalation bump watcher running")

    yield

    # --- SHUTDOWN ---
    for t in [alert_worker_task, notifier_task, redis_sub_task, denial_worker_task,esc_bump_task]:
        if t:
            t.cancel()
            try:
                await t
            except asyncio.CancelledError:
                pass

    if redis_client:
        await redis_client.aclose()
    if database.is_connected:
        await database.disconnect()
    if database2.is_connected:
        await database2.disconnect()

    logger.info("Application STOPPED")


# ============================================================
# FastAPI App Creation (MUST be before any @app routes)
# ============================================================
app = FastAPI(lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in ALLOWED_ORIGINS],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

router = APIRouter()
app.include_router(router)


# ============================================================
# ALL API ROUTES & WEBSOCKETS BELOW
# ============================================================

# -------------------- Login API --------------------
@app.post("/login")
async def login_user(data: LoginRequest):
    username = data.username
    password = data.password
    password_md5 = hashlib.md5(password.encode()).hexdigest()

    query = """
        SELECT clg_group, clg_is_login
        FROM ems_colleague
        WHERE clg_ref_id = :username AND clg_password = :password
    """
    result = await database.fetch_one(query, {"username": username, "password": password_md5})

    if not result:
        raise HTTPException(status_code=401, detail="Invalid username or password")

    update_query = """
        UPDATE ems_colleague SET clg_is_login = 'yes' WHERE clg_ref_id = :username
    """
    await database.execute(update_query, {"username": username})
    token = generate_token(username)

    return {"message": "Login successful", "status": "success", "username": username, "token": token}


# -------------------- Logout API --------------------
@app.post("/logout")
async def logout_user(user_id: str = Depends(verify_token)):
    update_query = """
        UPDATE ems_colleague
        SET clg_is_login = 'no'
        WHERE clg_ref_id = :username
    """
    await database.execute(update_query, {"username": user_id})

    return {"message": "Logout successful", "status": "success"}


# -------------------- Districts API --------------------
@app.get("/api/districts", response_model=List[DistrictOut])
async def get_districts(division_id: Optional[int] = Query(None)):
    if division_id is not None:
        query = """
            SELECT dst_code, dst_name
            FROM ems_mas_districts
            WHERE div_id = :division_id
            AND dst_state='MH'
            AND dstis_deleted = '0'
            ORDER BY dst_name
        """
        rows = await cached_query(query, {"division_id": division_id}, ttl=10)
    else:
        query = """
            SELECT dst_code, dst_name
            FROM ems_mas_districts
            WHERE dst_state='MH'
            AND dstis_deleted = '0'
            ORDER BY dst_name
        """
        rows = await cached_query(query, ttl=10)

    districts = [
        DistrictOut(district_id=row["dst_code"], district_name=row["dst_name"])
        for row in rows
    ]
    return districts


# -------------------- Division API --------------------
@app.get("/api/division", response_model=List[DivisionOut])
async def get_divisions():
    query = """
        SELECT div_code, div_name
        FROM ems_mas_division
        ORDER BY div_name
    """
    rows = await database.fetch_all(query)

    division = [
        DivisionOut(division_id=row["div_code"], division_name=row["div_name"])
        for row in rows
    ]
    return division


# -------------------- Ambulance List API --------------------
@app.get("/api/ambulance-list")
async def get_ambulance_list():
    query = """
        SELECT
            a.amb_rto_register_no AS ambulance_number,
            d.dst_name AS district_name,
            a.amb_district AS dst_code
        FROM ems_ambulance a
        LEFT JOIN ems_mas_districts d
            ON a.amb_district = d.dst_code
        WHERE a.amb_rto_register_no IS NOT NULL
        ORDER BY d.dst_name, a.amb_rto_register_no;
    """

    rows = await cached_query(query, ttl=30, fetch="all")
    if not rows:
        return {"status": "error", "message": "No ambulances found"}

    result = [
        {
            "ambulance_number": row["ambulance_number"],
            "district_name": row["district_name"],
            "dst_code": row["dst_code"]
        }
        for row in rows
    ]

    return {"status": "success", "count": len(result), "data": result}


# -------------------- RTM Dashboard WebSocket --------------------
@app.websocket("/ws/rtm_dashboard")
async def rtm_dashboard_ws(websocket: WebSocket):
    user_id = await verify_jwt_token(websocket.query_params.get("token"))
    if not user_id:
        logger.warning("RTM Dashboard WS rejected: invalid or missing token")
        await websocket.close(code=1008)
        return
    await websocket.accept()

    prev_data = None
    last_filter = {}
    last_filter_hash = None

    try:
        while True:
            try:
                msg = await asyncio.wait_for(
                    websocket.receive_json(), timeout=0.5
                )
                if msg:
                    last_filter = msg
            except asyncio.TimeoutError:
                pass

            if not last_filter:
                await asyncio.sleep(0.2)
                continue

            inc_ref_id = last_filter.get("inc_ref_id")
            if not inc_ref_id:
                continue

            current_filter_hash = hash_filter(last_filter)
            if current_filter_hash == last_filter_hash:
                await asyncio.sleep(0.2)
                continue

            last_filter_hash = current_filter_hash

            where_sql = "WHERE 1=1"
            params = {}

            where_sql += " AND inc_ref_id = :inc_ref_id"
            params["inc_ref_id"] = inc_ref_id

            if last_filter.get("dst_code"):
                where_sql += " AND dst_code = :dst_code"
                params["dst_code"] = last_filter["dst_code"]

            if last_filter.get("ambulance_no"):
                where_sql += " AND ambulance_no = :ambulance_no"
                params["ambulance_no"] = last_filter["ambulance_no"]

            query = f"""
                SELECT
                    inc_ref_id,
                    ambulance_no,
                    dst_code,
                    district_name,
                    base_location_name,
                    call_type,
                    caller_mobile,
                    pilot_name,
                    pilot_mobile,
                    paramedic_name,
                    paramedic_mobile,
                    assigned_time,
                    parameter_count,
                    inc_dispatch_time,
                    inc_recive_time,
                    inc_datetime,
                    acknowledge,
                    start_from_base_loc,
                    acknowledge_duration,
                    start_from_base_duration,
                    at_scene,
                    at_scene_duration,
                    wait_time_at_scene_duration,
                    from_scene,
                    start_from_scene_duration,
                    enroute_to_hospital_duration,
                    at_hospital,
                    at_hospital_duration,
                    patient_handover,
                    handover_duration,
                    back_to_base_loc,
                    back_to_base_duration,
                    inc_pcr_status,
                    clg_is_login,
                    destination_hospital_id,
                    rec_hospital_name,
                    hospital_id,
                    amb_working_area,
                    pilot_parameters,
                    is_validate,
                    trip,
                    remark,
                    pilot_login_out,
                    emso_login_out,
                    amb_type
                FROM rtm_dashboard
                {where_sql}
                ORDER BY inc_datetime::timestamp DESC
                LIMIT 1
            """

            rows = await cached_query(
                query,
                params=params,
                fetch="all",
                ttl=2,
                db=database2
            )

            data = [normalize_row(r) for r in rows] if rows else []

            if data != prev_data:
                await websocket.send_json({"latest_records": data})
                prev_data = data

    except WebSocketDisconnect:
        logger.info("RTM Dashboard WebSocket disconnected")


# -------------------- RTM Alerts WebSocket --------------------
@app.websocket("/ws/rtm_alerts")
async def rtm_alerts_ws(websocket: WebSocket):
    await websocket.accept()
    prev_alerts = None

    try:
        while True:
            query = """
                SELECT *
                FROM rtm_dashboard
                WHERE EXTRACT(YEAR FROM inc_datetime) = 2026
                ORDER BY inc_datetime DESC
                LIMIT 200
            """

            rows = await cached_query(
                query,
                fetch="all",
                ttl=5,
                db=database2
            )

            alerts = []

            for row in rows:
                row = normalize_row(row)

                inc_dispatch_sec = hhmmss_to_seconds(row.get("inc_dispatch_time"))
                acknowledge_sec = hhmmss_to_seconds(row.get("acknowledge_duration"))
                start_base_sec = hhmmss_to_seconds(row.get("start_from_base_duration"))
                at_scene_sec = hhmmss_to_seconds(row.get("at_scene_duration"))

                amb_area = row.get("amb_working_area")

                is_alert = False

                if inc_dispatch_sec > 150:
                    is_alert = True
                elif acknowledge_sec > 30:
                    is_alert = True
                elif start_base_sec > 120:
                    is_alert = True
                elif amb_area == "1" and at_scene_sec > 1500:
                    is_alert = True
                elif amb_area == "2" and at_scene_sec > 1080:
                    is_alert = True

                if is_alert:
                    alerts.append(row)

            if alerts != prev_alerts:
                await websocket.send_json({
                    "year": 2026,
                    "alert_count": len(alerts),
                    "alerts": alerts
                })
                prev_alerts = alerts

            await asyncio.sleep(3)

    except WebSocketDisconnect:
        logger.info("RTM Alert WebSocket disconnected")


# -------------------- Central Alerts WebSocket --------------------

# ---- tuning knobs ----
MAX_CONSECUTIVE_RECV_ERRORS = 5  # only bail if the socket is TRULY broken (not for one bad message)
RECV_ERROR_WINDOW = 10           # errors must cluster within this many seconds to count as "truly broken"

def _bump_error(consecutive_errors: int, first_error_ts):
    """Track error bursts so a single stray error never kills a healthy connection,
    but a genuinely dead socket (many errors back-to-back) still gets cleaned up."""
    now = time.monotonic()
    if first_error_ts is None or now - first_error_ts > RECV_ERROR_WINDOW:
        return 1, now
    return consecutive_errors + 1, first_error_ts

@app.websocket("/ws/central_alerts")
async def central_alerts_ws(websocket: WebSocket):

    user_id = await verify_jwt_token(websocket.query_params.get("token"))
    if not user_id:
        await websocket.accept()
        await websocket.send_json({
            "type": "ERROR",
            "status": 401,
            "message": "Invalid or expired token. Please login again."
        })
        await websocket.close(code=1008)
        return

    conn_info = await manager.connect(websocket, user_id)
    queue = conn_info["queue"]

    print("🔌 WebSocket client connected")

    try:
        # =====================================================
        # 1️⃣ SEND DATA ON CONNECT (TODAY DEFAULT)
        # =====================================================
        rows = await cached_query(
            """
            SELECT c.*, af.status AS alert_flow_status
            FROM central_alerts c
            LEFT JOIN LATERAL (
                SELECT status
                FROM alert_escalation_flow
                WHERE alert_id = c.alert_id
                LIMIT 1
            ) af ON TRUE
            WHERE c.inc_datetime >= CURRENT_DATE
              AND c.inc_datetime < CURRENT_DATE + INTERVAL '1 day'
            ORDER BY c.inc_datetime DESC, c.alert_id DESC
            """,
            ttl=3,
            fetch="all",
            db=database2
        )

        today_all = [serialize_row(r) for r in rows]
        by_severity = group_by_severity(rows)

        # ===============================
        # COUNTS (108 / 102)
        # ===============================
        count_rows = await cached_query(
            """
            SELECT
                system_type,
                severity,
                COUNT(*) AS total
            FROM central_alerts
            WHERE inc_datetime >= CURRENT_DATE
              AND inc_datetime < CURRENT_DATE + INTERVAL '1 day'
              AND escalate_status = '1'
              AND system_type IN ('108', '102')
            GROUP BY system_type, severity
            """,
            ttl=3,
            fetch="all",
            db=database2
        )

        counts = {
            "total": {"108": 0, "102": 0},
            "severity": {"108": {}, "102": {}}
        }

        for r in count_rows:
            system = r["system_type"]
            severity = r["severity"]
            total = r["total"]

            counts["total"][system] += total
            counts["severity"][system][severity] = total

        await websocket.send_json({
            "type": "ALL_ALERTS",
            "data": {
                "today_all": today_all,
                "by_severity": by_severity,
                "counts": counts
            }
        })

        print("📤 Sent ALL_ALERTS on connect")

        # =====================================================
        # 2️⃣ TWO CONCURRENT LOOPS — real-time updates ke liye
        # =====================================================

        # 👇 YE FUNCTION AB BATCH QUERY MAREGA (100% SAFE & FAST)
        async def ensure_status(records):
            if not records:
                return records
                
            # 1. Un records ko nikaalo jisme status missing hai
            missing_records = [r for r in records if "alert_flow_status" not in r or r.get("alert_flow_status") is None]
            
            # Agar sabke paas status hai, toh kuch karne ki zaroorat nahi
            if not missing_records:
                return records
                
            # 2. Missing records ke alert_ids nikaalo
            alert_ids = [r.get("alert_id") for r in missing_records if r.get("alert_id") is not None]
            
            if not alert_ids:
                for r in missing_records:
                    r["alert_flow_status"] = None
                return records
                
            try:
                # 3. Data type issue fix karo
                try:
                    clean_ids = [int(aid) for aid in alert_ids]
                except (ValueError, TypeError):
                    clean_ids = alert_ids
                    
                # 4. Ek hi query mein saare IDs ka status nikaalo (Batch Query)
                ids_str = ",".join(str(aid) for aid in clean_ids)
                q = f"""
                    SELECT alert_id, status 
                    FROM alert_escalation_flow 
                    WHERE alert_id IN ({ids_str})
                """
                rows = await database2.fetch_all(q)
                
                # 5. Dictionary bana lo taaki map karna fast ho
                status_map = {}
                for row in rows:
                    try:
                        status_map[int(row["alert_id"])] = row["status"]
                    except (ValueError, TypeError):
                        status_map[str(row["alert_id"])] = row["status"]
                        
                # 6. Missing records mein status attach karo
                for r in missing_records:
                    aid = r.get("alert_id")
                    status = None
                    try:
                        status = status_map.get(int(aid))
                    except (ValueError, TypeError):
                        status = status_map.get(str(aid))
                    r["alert_flow_status"] = status
                    
            except Exception as e:
                print(f"❌ Error in ensure_status: {e}")
                # Agar query fail ho jaye, toh None set kar do, disconnect mat karo
                for r in missing_records:
                    r["alert_flow_status"] = None
                
            return records


        async def drain_loop():
            """Queue se broadcast messages nikal ke client pe bhejo."""
            while True:
                payload = await queue.get()
                msg_type = payload.get("type")

                if msg_type == "ALL_ALERTS":
                    data = payload.get("data", {})
                    records = data.get("today_all", [])
                    
                    # 👇 YAHAN ENSURE_STATUS CALL HO RAHA HAI
                    data["today_all"] = await ensure_status(records)

                    if not await manager.safe_send(websocket, payload):
                        break

        async def client_listener():
            """Client se filter requests receive karo."""
            while True:
                msg = await websocket.receive_json()

                incident_id = msg.get("incident_id")
                filter_date = msg.get("date")  # YYYY-MM-DD

                conditions = []
                params = {}

                if incident_id:
                    conditions.append("c.incident_id = :incident_id")
                    params["incident_id"] = str(incident_id)

                if filter_date:
                    start_dt = datetime.strptime(filter_date, "%Y-%m-%d")
                    end_dt = start_dt + timedelta(days=1)

                    conditions.append("""
                        c.created_date >= :start_date
                        AND c.created_date < :end_date
                    """)

                    params["start_date"] = start_dt
                    params["end_date"] = end_dt
                else:
                    conditions.append("""
                        c.inc_datetime >= CURRENT_DATE
                        AND c.inc_datetime < CURRENT_DATE + INTERVAL '1 day'
                    """)

                where_clause = " AND ".join(conditions)

                query = f"""
                    SELECT c.*, af.status AS alert_flow_status
                    FROM central_alerts c
                    LEFT JOIN LATERAL (
                        SELECT status
                        FROM alert_escalation_flow
                        WHERE alert_id = c.alert_id
                        LIMIT 1
                    ) af ON TRUE
                    WHERE {where_clause}
                    ORDER BY c.inc_datetime DESC, c.alert_id DESC
                """

                if filter_date:
                    rows = await database2.fetch_all(query, params)
                else:
                    rows = await cached_query(
                        query,
                        params=params,
                        ttl=3,
                        fetch="all",
                        db=database2
                    )

                today_all = [serialize_row(r) for r in rows]
                
                # 👇 YAHAN BHI ENSURE_STATUS CALL HO RAHA HAI
                today_all = await ensure_status(today_all)
                
                by_severity = group_by_severity(rows)

                await manager.safe_send(websocket, {
                    "type": "ALL_ALERTS",
                    "data": {
                        "today_all": today_all,
                        "by_severity": by_severity,
                        "counts": counts
                    }
                })

                print(f"📤 Sent alerts (incident_id={incident_id}, date={filter_date})")

        # 👇 Dono loops concurrently chalao
        sender = asyncio.create_task(drain_loop())
        receiver = asyncio.create_task(client_listener())

        done, pending = await asyncio.wait(
            [sender, receiver], return_when=asyncio.FIRST_COMPLETED
        )
        for t in pending:
            t.cancel()
        for t in done:
            if t.exception():
                logger.exception(f"central_alerts WS task FAILED: {t.exception()}")

    except WebSocketDisconnect:
        print("❌ WebSocket client disconnected")
    except Exception as e:
        print(f"❌ WebSocket Error: {e}")
    finally:
        manager.disconnect(websocket)

# -------------------- Escalate API --------------------
@app.put("/api/escalate/{alert_id}")
async def escalate_alert(alert_id: int, payload: EscalateRequest):
    """Escalate alert: escalate_status 1 -> 2, update remark and timestamps."""
    alert = await database2.fetch_one(
        """
        SELECT alert_id, escalate_status
        FROM central_alerts
        WHERE alert_id = :alert_id
        """,
        {"alert_id": alert_id}
    )

    if not alert:
        raise HTTPException(status_code=404, detail="Alert not found")

    await database2.execute(
        """
        UPDATE central_alerts
        SET escalate_status = 2,
            escalated_deny_remark = :remark,
            updated_date = NOW(),
            escalated_date = NOW(),
            escalated_by = :escalated_by
        WHERE alert_id = :alert_id
        """,
        {
            "alert_id": alert_id,
            "remark": payload.remark,
            "escalated_by": payload.escalated_by
        }
    )

    return {
        "status": "success",
        "message": "Alert escalated successfully",
        "alert_id": alert_id,
        "escalate_status": 2,
        "remark": payload.remark,
        "updated_date": datetime.now(timezone.utc).isoformat()
    }


# -------------------- Dashboard Overview API --------------------
# @app.get("/api/dashboard")
# async def dashboard_alerts_overview(
#     range: Optional[str] = Query("today", enum=["today", "month", "all"])
# ):
#     date_filter = get_date_filter(range)

#     total_alerts_sql = f"""
#     SELECT
#         COUNT(*) as total,
#         COUNT(*) FILTER (WHERE system_type='108') as system_108,
#         COUNT(*) FILTER (WHERE system_type='102') as system_102
#     FROM public.central_alerts
#     WHERE is_deleted = false
#     AND {date_filter}
#     """
#     total_alerts = await database2.fetch_one(total_alerts_sql)

#     escalated_sql = f"""
#     SELECT
#         COUNT(*) as total,
#         COUNT(*) FILTER (WHERE system_type='108') as system_108,
#         COUNT(*) FILTER (WHERE system_type='102') as system_102
#     FROM public.central_alerts
#     WHERE is_deleted = false
#     AND escalate_status = '2'
#     AND {date_filter}
#     """
#     escalated_alerts = await database2.fetch_one(escalated_sql)

#     severity_sql = f"""
#     SELECT
#         severity,
#         COUNT(*) FILTER (WHERE system_type='108') as system_108,
#         COUNT(*) FILTER (WHERE system_type='102') as system_102
#     FROM public.central_alerts
#     WHERE is_deleted = false
#     AND {date_filter}
#     GROUP BY severity
#     ORDER BY severity
#     """
#     severity_rows = await database2.fetch_all(severity_sql)

#     severity_data = [
#         {
#             "severity": row["severity"],
#             "system_108": row["system_108"],
#             "system_102": row["system_102"]
#         }
#         for row in severity_rows
#     ]

#     return {
#         "total_alerts": dict(total_alerts),
#         "escalated_alerts": dict(escalated_alerts),
#         "severity_timeline": severity_data
#     }

@app.get("/api/dashboard")
async def dashboard_alerts_overview(
    range: Optional[str] = Query("today", enum=["today", "month", "all"])
):
    date_filter = get_date_filter(range)
    # Denial table ke liye date filter banayein (kyunki uska column naam 'added_date' hai)
    if range == "today":
        denial_date_filter = "DATE(added_date) = CURRENT_DATE"
    elif range == "month":
        denial_date_filter = "DATE_TRUNC('month', added_date) = DATE_TRUNC('month', CURRENT_DATE)"
    else:
        denial_date_filter = "1=1"
 
    # --------------------------------------------------------
    # 1. CENTRAL ALERTS (Pehle jaisa)
    # --------------------------------------------------------
    total_alerts_sql = f"""
    SELECT
        COUNT(*) as total,
        COUNT(*) FILTER (WHERE system_type='108') as system_108,
        COUNT(*) FILTER (WHERE system_type='102') as system_102
    FROM public.central_alerts
    WHERE is_deleted = false
    AND {date_filter}
    """
    total_alerts = await database2.fetch_one(total_alerts_sql)
    total_alerts = dict(total_alerts) if total_alerts else {"total": 0, "system_108": 0, "system_102": 0}
 
    escalated_sql = f"""
    SELECT
        COUNT(*) as total,
        COUNT(*) FILTER (WHERE system_type='108') as system_108,
        COUNT(*) FILTER (WHERE system_type='102') as system_102
    FROM public.central_alerts
    WHERE is_deleted = false
    AND escalate_status = '2'
    AND {date_filter}
    """
    escalated_alerts = await database2.fetch_one(escalated_sql)
    escalated_alerts = dict(escalated_alerts) if escalated_alerts else {"total": 0, "system_108": 0, "system_102": 0}
 
    severity_sql = f"""
    SELECT
        severity,
        COUNT(*) FILTER (WHERE system_type='108') as system_108,
        COUNT(*) FILTER (WHERE system_type='102') as system_102
    FROM public.central_alerts
    WHERE is_deleted = false
    AND {date_filter}
    GROUP BY severity
    ORDER BY severity
    """
    severity_rows = await database2.fetch_all(severity_sql)
 
    severity_data = [
        {
            "severity": row["severity"],
            "system_108": row["system_108"],
            "system_102": row["system_102"]
        }
        for row in severity_rows
    ]
 
    # --------------------------------------------------------
    # 2. DENIAL RECORDS (NAYA: Distinct call_id count karega)
    # --------------------------------------------------------
    # Total Denials (Distinct call_id)
    denial_total_sql = f"""
    SELECT COUNT(DISTINCT call_id) as total
    FROM public.denial_escalation_master
    WHERE (is_deleted = FALSE OR is_deleted IS NULL)
    AND {denial_date_filter}
    """
    denial_total = await database2.fetch_one(denial_total_sql)
    denial_count = denial_total["total"] or 0 if denial_total else 0
 
    # Escalated Denials (Distinct call_id where status = 2)
    denial_escalated_sql = f"""
    SELECT COUNT(DISTINCT call_id) as total
    FROM public.denial_escalation_master
    WHERE (is_deleted = FALSE OR is_deleted IS NULL)
    AND escalate_status = '2'
    AND {denial_date_filter}
    """
    denial_escalated = await database2.fetch_one(denial_escalated_sql)
    denial_escalated_count = denial_escalated["total"] or 0 if denial_escalated else 0
 
    # --------------------------------------------------------
    # 3. FINAL COMBINED DATA (Dono table ka sum)
    # --------------------------------------------------------
    final_total = {
        "total": total_alerts["total"] + denial_count,
        "system_108": total_alerts["system_108"] + denial_count,  # 👈 Denial 108 me add hoga
        "system_102": total_alerts["system_102"]
    }
 
    final_escalated = {
        "total": escalated_alerts["total"] + denial_escalated_count,
        "system_108": escalated_alerts["system_108"] + denial_escalated_count, # 👈 Escalated denial 108 me add hoga
        "system_102": escalated_alerts["system_102"]
    }
 
    return {
        "total_alerts": final_total,
        "escalated_alerts": final_escalated,
        "severity_timeline": severity_data
    }


# -------------------- Excel Report Download API --------------------
# @app.get("/api/dashboard/download-client-report")
# async def download_client_report(
#     range_type: str = Query("today", enum=["today", "month", "all"])
# ):
#     date_filter = get_date_filter(range_type)

#     full_sql = f"""
#     SELECT *
#     FROM public.central_alerts
#     WHERE is_deleted = false
#     AND {date_filter}
#     ORDER BY created_date DESC
#     """
#     rows = await database2.fetch_all(full_sql)

#     df = pd.DataFrame([dict(r) for r in rows]) if rows else pd.DataFrame()

#     if not df.empty:
#         df.insert(0, "Sr No", list(range(1, len(df) + 1)))

#         escalation_map = {
#             "0": "Open",
#             "1": "In Progress",
#             "2": "Escalated",
#             "3": "Closed"
#         }

#         if "escalate_status" in df.columns:
#             df["Escalation Status"] = df["escalate_status"].astype(str).map(escalation_map)
#             df.drop(columns=["escalate_status"], inplace=True)

#         if "is_deleted" in df.columns:
#             df.drop(columns=["is_deleted"], inplace=True)

#         date_cols = ["created_date", "updated_date", "cancel_date", "escalated_date"]
#         for col in date_cols:
#             if col in df.columns:
#                 df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%d-%m-%Y %H:%M")

#         rename_map = {
#             "severity": "Severity",
#             "created_date": "Created Date & Time",
#             "updated_date": "Updated Date & Time",
#             "division": "Division",
#             "district": "District",
#             "inc_latitude": "Incidence Latitude",
#             "inc_longitude": "Incidence Longitude",
#             "amb_lat": "Ambulance Lattitude",
#             "amb_long": "Ambulance Longitude",
#             "paramedic_name": "EMT Name",
#             "paramedic_mobile": "EMT Mobile",
#             "inc_datetime": "Incidence Datetime",
#             "alert_type": "Alert Type",
#             "incident_id": "Incident Id",
#             "ambulance_no": "Ambulance Number",
#             "remark": "Remark",
#             "escalated_deny_remark": "Escalated/Deny Remark",
#             "pilot_name": "Pilot Name",
#             "pilot_mobile": "Pilot Mobile",
#             "escalated_date": "Escalated Date",
#             "cancel_date": "Cancel Date",
#             "escalated_by": "Escalated By",
#             "cancel_by": "Cancel By",
#             "Escalation Status": "Escalation Status",
#             "system_type": "System Type",
#             "alert_id": "Alert ID",
#             "Sr No": "Sr No",
#         }

#         df.rename(columns=rename_map, inplace=True)

#         column_order = [
#             "Sr No", "Alert ID", "Alert Type", "System Type", "Severity",
#             "Incident Id", "Incidence Datetime", "Division", "District",
#             "Incidence Latitude", "Incidence Longitude", "Ambulance Number",
#             "Ambulance Lattitude", "Ambulance Longitude", "Pilot Name",
#             "Pilot Mobile", "EMT Name", "EMT Mobile",
#             "Created Date & Time", "Updated Date & Time",
#             "Escalation Status", "Escalated Date", "Escalated By",
#             "Escalated/Deny Remark", "Cancel Date", "Cancel By", "Remark",
#         ]

#         existing_ordered_cols = [c for c in column_order if c in df.columns]
#         remaining_cols = [c for c in df.columns if c not in existing_ordered_cols]
#         df = df[existing_ordered_cols + remaining_cols]
#     else:
#         df = pd.DataFrame([{"Message": "No Data Found"}])

#     summary_sql = f"""
#     SELECT
#         COUNT(*) as total_alerts,
#         COUNT(*) FILTER (WHERE escalate_status='2') as escalated_alerts,
#         COUNT(*) FILTER (WHERE system_type='108') as system_108
#     FROM public.central_alerts
#     WHERE is_deleted = false
#     AND {date_filter}
#     """
#     summary_row = await database2.fetch_one(summary_sql)
#     df_summary = pd.DataFrame([dict(summary_row)]) if summary_row else pd.DataFrame()

#     summary_rename_map = {
#         "total_alerts": "Total Alerts",
#         "escalated_alerts": "Escalated Alerts",
#         "system_108": "System 108"
#     }
#     if not df_summary.empty:
#         df_summary.rename(columns=summary_rename_map, inplace=True)

#     output = io.BytesIO()
#     with pd.ExcelWriter(output, engine="openpyxl") as writer:
#         df.to_excel(writer, sheet_name="All Alert Records", index=False)
#         df_summary.to_excel(writer, sheet_name="Summary", index=False)
#         workbook = writer.book
#         format_worksheet(workbook["All Alert Records"])
#         format_worksheet(workbook["Summary"])

#     output.seek(0)
#     file_name = f"Central_Alerts_Client_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

#     return StreamingResponse(
#         output,
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         headers={"Content-Disposition": f"attachment; filename={file_name}"}
#     )

@app.get("/api/dashboard/download-client-report")

async def download_client_report(

    range_type: str = Query("today", enum=["today", "month", "all"])

):

    # 1. Central Alerts ke liye date filter

    date_filter = get_date_filter(range_type)

    # 2. Denial Alerts ke liye alag date filter (kyunki column naam 'added_date' hai)

    if range_type == "today":

        denial_date_filter = "DATE(added_date) = CURRENT_DATE"

    elif range_type == "month":

        denial_date_filter = "DATE_TRUNC('month', added_date) = DATE_TRUNC('month', CURRENT_DATE)"

    else:

        denial_date_filter = "1=1"
 
    # --------------------------------------------------------

    # 1. FETCH CENTRAL ALERTS

    # --------------------------------------------------------

    full_sql = f"""

    SELECT *

    FROM public.central_alerts

    WHERE is_deleted = false

    AND {date_filter}

    ORDER BY created_date DESC

    """

    rows = await database2.fetch_all(full_sql)
 
    df = pd.DataFrame([dict(r) for r in rows]) if rows else pd.DataFrame()
 
    if not df.empty:

        df.insert(0, "Sr No", list(range(1, len(df) + 1)))
 
        escalation_map = {

            "0": "Open",

            "1": "In Progress",

            "2": "Escalated",

            "3": "Closed"

        }
 
        if "escalate_status" in df.columns:

            df["Escalation Status"] = df["escalate_status"].astype(str).map(escalation_map)

            df.drop(columns=["escalate_status"], inplace=True)
 
        if "is_deleted" in df.columns:

            df.drop(columns=["is_deleted"], inplace=True)
 
        date_cols = ["created_date", "updated_date", "cancel_date", "escalated_date"]

        for col in date_cols:

            if col in df.columns:

                df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%d-%m-%Y %H:%M")
 
        rename_map = {

            "severity": "Severity",

            "created_date": "Created Date & Time",

            "updated_date": "Updated Date & Time",

            "division": "Division",

            "district": "District",

            "inc_latitude": "Incidence Latitude",

            "inc_longitude": "Incidence Longitude",

            "amb_lat": "Ambulance Lattitude",

            "amb_long": "Ambulance Longitude",

            "paramedic_name": "EMT Name",

            "paramedic_mobile": "EMT Mobile",

            "inc_datetime": "Incidence Datetime",

            "alert_type": "Alert Type",

            "incident_id": "Incident Id",

            "ambulance_no": "Ambulance Number",

            "remark": "Remark",

            "escalated_deny_remark": "Escalated/Deny Remark",

            "pilot_name": "Pilot Name",

            "pilot_mobile": "Pilot Mobile",

            "escalated_date": "Escalated Date",

            "cancel_date": "Cancel Date",

            "escalated_by": "Escalated By",

            "cancel_by": "Cancel By",

            "Escalation Status": "Escalation Status",

            "system_type": "System Type",

            "alert_id": "Alert ID",

            "Sr No": "Sr No",

        }
 
        df.rename(columns=rename_map, inplace=True)
 
        column_order = [

            "Sr No", "Alert ID", "Alert Type", "System Type", "Severity",

            "Incident Id", "Incidence Datetime", "Division", "District",

            "Incidence Latitude", "Incidence Longitude", "Ambulance Number",

            "Ambulance Lattitude", "Ambulance Longitude", "Pilot Name",

            "Pilot Mobile", "EMT Name", "EMT Mobile",

            "Created Date & Time", "Updated Date & Time",

            "Escalation Status", "Escalated Date", "Escalated By",

            "Escalated/Deny Remark", "Cancel Date", "Cancel By", "Remark",

        ]
 
        existing_ordered_cols = [c for c in column_order if c in df.columns]

        remaining_cols = [c for c in df.columns if c not in existing_ordered_cols]

        df = df[existing_ordered_cols + remaining_cols]

    else:

        df = pd.DataFrame([{"Message": "No Incident Alert Data Found"}])
 
 
     # --------------------------------------------------------
    # 2. FETCH DENIAL RECORDS (NAYA SHEET KE LIYE)
    # --------------------------------------------------------
    denial_sql = f"""
    SELECT id, mysql_id, call_id, amb_no, amb_default_mobile, caller_no,
           hp_name, challenge_val, meaning, denial_remark, remark, dst_name,
           alert_type, added_by, added_date, escalate_status
    FROM public.denial_escalation_master
    WHERE (is_deleted = FALSE OR is_deleted IS NULL)
    AND {denial_date_filter}
    ORDER BY added_date DESC
    """
 
    denial_rows = await database2.fetch_all(denial_sql)
    df_denial = pd.DataFrame([dict(r) for r in denial_rows]) if denial_rows else pd.DataFrame()
 
    if not df_denial.empty:
        df_denial.insert(0, "Sr No", list(range(1, len(df_denial) + 1)))
 
        # Denial ke liye Escalation Map
        denial_escalation_map = {
            "0": "Open",
            "1": "In Progress",
            "2": "Escalated",
            "3": "Closed"
        }
 
        if "escalate_status" in df_denial.columns:
            df_denial["Escalation Status"] = df_denial["escalate_status"].astype(str).map(denial_escalation_map)
            df_denial.drop(columns=["escalate_status"], inplace=True)
 
        if "added_date" in df_denial.columns:
            df_denial["Added Date & Time"] = pd.to_datetime(df_denial["added_date"], errors="coerce").dt.strftime("%d-%m-%Y %H:%M")
            df_denial.drop(columns=["added_date"], inplace=True)
 
        if "mysql_id" in df_denial.columns:
            df_denial.drop(columns=["mysql_id"], inplace=True)
 
        denial_rename_map = {
            "Sr No": "Sr No",
            "id": "Denial ID",
            "call_id": "Call ID",
            "amb_no": "Ambulance Number",
            "amb_default_mobile": "Ambulance Mobile",
            "caller_no": "Caller No",
            "hp_name": "Hospital Name",
            "challenge_val": "Challenge",
            "meaning": "Denial Reason",
            "denial_remark": "Denial Remark",
            "remark": "Remark",                        # ✅ NAYA
            "dst_name": "District",
            "alert_type": "Alert Type",
            "added_by": "Added By",
            "Added Date & Time": "Added Date & Time",
            "Escalation Status": "Escalation Status"
        }
 
        df_denial.rename(columns=denial_rename_map, inplace=True)
 
        # ✅ NAYA: Column order — Remark ko Escalation Status ke right me rakhne ke liye
        denial_column_order = [
            "Sr No", "Denial ID", "Call ID", "Ambulance Number", "Ambulance Mobile",
            "Caller No", "Hospital Name", "Challenge", "Denial Reason",
            "Denial Remark", "District", "Alert Type", "Added By",
            "Added Date & Time", "Escalation Status", "Remark",
        ]
 
        denial_ordered_cols = [c for c in denial_column_order if c in df_denial.columns]
        denial_remaining_cols = [c for c in df_denial.columns if c not in denial_ordered_cols]
        df_denial = df_denial[denial_ordered_cols + denial_remaining_cols]
 
    else:
        df_denial = pd.DataFrame([{"Message": "No Call Denial Data Found"}])
 
 
    # --------------------------------------------------------

    # 3. COMBINED SUMMARY CALCULATIONS

    # --------------------------------------------------------

    # Central Counts

    summary_sql = f"""

    SELECT

        COUNT(*) as total_alerts,

        COUNT(*) FILTER (WHERE escalate_status='2') as escalated_alerts,

        COUNT(*) FILTER (WHERE system_type='108') as system_108

    FROM public.central_alerts

    WHERE is_deleted = false

    AND {date_filter}

    """

    summary_row = await database2.fetch_one(summary_sql)

    summary_data = dict(summary_row) if summary_row else {"total_alerts": 0, "escalated_alerts": 0, "system_108": 0}
 
    # Denial Counts (Distinct call_id)

    denial_summary_sql = f"""

    SELECT

        COUNT(DISTINCT call_id) as denial_total,

        COUNT(DISTINCT call_id) FILTER (WHERE escalate_status='2') as denial_escalated

    FROM public.denial_escalation_master

    WHERE (is_deleted = FALSE OR is_deleted IS NULL)

    AND {denial_date_filter}

    """

    denial_summary_row = await database2.fetch_one(denial_summary_sql)

    denial_summary_data = dict(denial_summary_row) if denial_summary_row else {"denial_total": 0, "denial_escalated": 0}
 
    # Final Summary DataFrame

    final_summary = {

        "Total Incident Alerts": summary_data.get("total_alerts", 0),

        "Total Call Denial Alerts (Distinct Call ID)": denial_summary_data.get("denial_total", 0),

        "Grand Total (Incident + Denial)": summary_data.get("total_alerts", 0) + denial_summary_data.get("denial_total", 0),

        "Escalated Incident Alerts": summary_data.get("escalated_alerts", 0),

        "Escalated Call Denial Alerts": denial_summary_data.get("denial_escalated", 0),

        "Total System 108 (Incident + Denial)": summary_data.get("system_108", 0) + denial_summary_data.get("denial_total", 0)

    }

    df_summary = pd.DataFrame([final_summary])
 
 
    # --------------------------------------------------------

    # 4. EXCEL GENERATION WITH MULTIPLE SHEETS

    # --------------------------------------------------------

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # 👇 Sheet 1 ka naam change kiya gaya hai

        df.to_excel(writer, sheet_name="Incident Alerts", index=False)

        # 👇 Sheet 2 ka naam change kiya gaya hai

        df_denial.to_excel(writer, sheet_name="Call Denial Alerts", index=False)

        df_summary.to_excel(writer, sheet_name="Summary", index=False)

        workbook = writer.book

        # 👇 Yahan bhi same naam use karna padega

        format_worksheet(workbook["Incident Alerts"])

        format_worksheet(workbook["Call Denial Alerts"])

        format_worksheet(workbook["Summary"])
 
    output.seek(0)

    file_name = f"Central_Alerts_Client_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
 
    return StreamingResponse(

        output,

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={"Content-Disposition": f"attachment; filename={file_name}"}

    )
 


# -------------------- Severity Update API --------------------
@app.put("/api/severity")
async def update_severity(data: SeverityUpdate):
    await database2.execute(
        """
        UPDATE central_alerts
        SET severity = :severity,
            updated_date = NOW()
        WHERE alert_id = :alert_id
        """,
        {"severity": data.severity, "alert_id": data.alert_id}
    )
    return {"message": "updated"}


# -------------------- Cancel Alert API --------------------
@app.put("/api/cancel")
async def cancel_alert(data: CancelUpdate):
    await database2.execute(
        """
        UPDATE central_alerts
        SET escalated_deny_remark = :remark,
            escalate_status = '3',
            updated_date = CURRENT_TIMESTAMP,
            cancel_date = CURRENT_TIMESTAMP,
            cancel_by = :cancel_by
        WHERE alert_id = :alert_id
        """,
        {
            "remark": data.remark,
            "alert_id": data.alert_id,
            "cancel_by": data.cancel_by
        }
    )
    return {"message": "Alert cancelled successfully"}


# -------------------- Top Ambulances WebSocket --------------------
@app.websocket("/ws/top-ambulances")
async def ws_top_ambulances(websocket: WebSocket):
    await websocket.accept()
    logger.info("Top Ambulances WebSocket connected")

    month = datetime.now().month
    lock = asyncio.Lock()
    last_payload = None

    async def fetch_and_send(force=False):
        nonlocal last_payload
        async with lock:
            query = """
                SELECT
                    ambulance_no,
                    CAST(SUM(total_alerts) AS BIGINT)        AS total_alerts,
                    CAST(SUM(start_delay) AS BIGINT)         AS start_delay_count,
                    CAST(SUM(at_scene_delay) AS BIGINT)      AS at_scene_delay_count,
                    CAST(SUM(ack_delay) AS BIGINT)           AS ack_delay_count,
                    CAST(SUM(mdt_not_logged_in) AS BIGINT)   AS mdt_not_logged_in_count,
                    CAST(SUM(back_to_base_delay) AS BIGINT)  AS back_to_base_delay_count
                FROM alerts_performance_monthly
                WHERE month = :month
                  AND ambulance_no IS NOT NULL
                GROUP BY ambulance_no
                ORDER BY total_alerts DESC
                LIMIT 50;
            """
            rows = await cached_query(
                query,
                params={"month": month},
                ttl=3,
                fetch="all",
                db=database2
            ) or []

            payload = {
                "month": month,
                "month_name": calendar.month_name[month],
                "top_ambulances": rows
            }

            if force or payload != last_payload:
                await websocket.send_json(payload)
                last_payload = payload

    async def listen_filters():
        nonlocal month
        while True:
            data = await websocket.receive_json()
            if "month" in data:
                m = int(data["month"])
                if 1 <= m <= 12 and m != month:
                    month = m
                    await fetch_and_send(force=True)

    async def auto_refresh():
        while True:
            await asyncio.sleep(3)
            await fetch_and_send()

    try:
        await fetch_and_send(force=True)
        await asyncio.gather(listen_filters(), auto_refresh())

    except WebSocketDisconnect:
        logger.info("Top Ambulances WebSocket disconnected")
    except Exception as e:
        logger.error(f"Top Ambulances WS error: {e}")
        await websocket.close()


# -------------------- Alert Thresholds API --------------------
@app.get("/api/alert-thresholds")
async def get_alert_thresholds():
    query = """
        SELECT *
        FROM alert_thresholds
        ORDER BY priority ASC;
    """
    rows = await cached_query(query, fetch="all", ttl=30, db=database2)
    return [dict(r) for r in rows]


@app.put("/api/update-alert-threshold/update/{id}")
async def update_alert_threshold(id: int, data: AlertThresholdUpdate):
    check_query = """
        SELECT id
        FROM alert_thresholds
        WHERE id = :id
    """
    row = await database2.fetch_one(query=check_query, values={"id": id})

    if not row:
        raise HTTPException(status_code=404, detail="Alert threshold not found.")

    update_fields = []
    values = {"id": id}

    if data.threshold_seconds is not None:
        update_fields.append("threshold_seconds = :threshold_seconds")
        values["threshold_seconds"] = data.threshold_seconds

    if data.severity is not None:
        update_fields.append("severity = :severity")
        values["severity"] = data.severity

    if data.priority is not None:
        update_fields.append("priority = :priority")
        values["priority"] = data.priority

    if not update_fields:
        raise HTTPException(
            status_code=400,
            detail="Please provide at least one field to update."
        )

    update_fields.append("updated_at = CURRENT_TIMESTAMP")

    query = f"""
        UPDATE alert_thresholds
        SET {', '.join(update_fields)}
        WHERE id = :id
    """

    try:
        await database2.execute(query=query, values=values)
        return {
            "status": "success",
            "message": "Alert threshold updated successfully.",
            "id": id
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# -------------------- Health Check Endpoint --------------------
@app.get("/health")
async def health_check():
    """Health check for load balancers and monitoring."""
    try:
        await database2.fetch_one("SELECT 1")
        db_status = "ok"
    except Exception:
        db_status = "error"

    try:
        r = await init_redis()
        await r.ping()
        redis_status = "ok"
    except Exception:
        redis_status = "error"

    notifier_running = notifier_task is not None and not notifier_task.done()
    worker_running = alert_worker_task is not None and not alert_worker_task.done()
    subscriber_running = redis_sub_task is not None and not redis_sub_task.done()

    all_ok = db_status == "ok" and notifier_running and worker_running

    return {
        "status": "healthy" if all_ok else "degraded",
        "database": db_status,
        "redis": redis_status,
        "notifier_running": notifier_running,
        "alert_worker_running": worker_running,
        "redis_subscriber_running": subscriber_running,
        "active_ws_connections": len(manager.active_connections),
        "timestamp": datetime.now(ist).isoformat()
    }

####################################################################################################
# def group_by_call_id(records: list[dict]) -> list[dict]:
#     """Same call_id wale records ko ek hi group/array mein daal deta hai."""
#     grouped = {}
#     for r in records:
#         cid = r.get("call_id")
#         grouped.setdefault(cid, []).append(r)

#     return [{"call_id": cid, "records": recs} for cid, recs in grouped.items()]


# @app.websocket("/ws/denial_alerts")
# async def websocket_escalation_alerts(websocket: WebSocket):
#     """Sirf naya record add hone par hit hoga."""

#     await websocket.accept()

#     active_ids = set()

#     try:
#         query = """
#             SELECT *
#             FROM denial_escalation_master
#             WHERE escalate_status = '1' 
#             AND (is_deleted = FALSE OR is_deleted IS NULL)
#             ORDER BY added_date DESC;
#         """

#         rows = await database2.fetch_all(query)
#         initial_data = []

#         for row in rows:
#             d = dict(row)
#             for k, v in d.items():
#                 if isinstance(v, datetime):
#                     d[k] = v.isoformat()
#             initial_data.append(d)
#             active_ids.add(d["id"])  # Current IDs track kar lo

#         if initial_data:
#             grouped_initial = group_by_call_id(initial_data)
#             await websocket.send_json({
#                 "type": "INITIAL_LOAD",
#                 "data": grouped_initial
#             })

#         # ---------------------------------------------------------
#         # STEP 2: Loop to check for NEW records
#         # ---------------------------------------------------------
#         while True:
#             await asyncio.sleep(5)  # Har 5 second mein DB check karo

#             current_rows = await database2.fetch_all(query)
#             current_data = []
#             current_ids = set()

#             for row in current_rows:
#                 d = dict(row)
#                 for k, v in d.items():
#                     if isinstance(v, datetime):
#                         d[k] = v.isoformat()
#                 current_data.append(d)
#                 current_ids.add(d["id"])

#             # Naye records find karo (jo active_ids mein nahi the)
#             new_records = [d for d in current_data if d["id"] not in active_ids]

#             # Agar naya record add hua hai tohi WS hit karo — call_id se group karke bhejo
#             if new_records:
#                 grouped_new = group_by_call_id(new_records)
#                 await websocket.send_json(grouped_new)

#             # Active IDs ko update karo
#             active_ids = current_ids

#     except WebSocketDisconnect:
#         print("Frontend disconnected from WebSocket")
#     except Exception as e:
#         print(f"WebSocket Error: {e}")
#         await websocket.close()


# def group_by_call_id(records: list[dict]) -> list[dict]:
#     """Same call_id wale records ko ek hi group/array mein daal deta hai,
#     aur added_date ke hisaab se sabse latest sabse upar rakhta hai."""

#     def parse_date(rec):
#         val = rec.get("added_date")
#         if isinstance(val, str):
#             try:
#                 return datetime.fromisoformat(val)
#             except (ValueError, TypeError):
#                 return datetime.min
#         elif isinstance(val, datetime):
#             return val
#         return datetime.min

#     grouped = {}
#     for r in records:
#         cid = r.get("call_id")
#         grouped.setdefault(cid, []).append(r)

#     result = []
#     for cid, recs in grouped.items():
#         recs_sorted = sorted(recs, key=parse_date, reverse=True)
#         result.append({"call_id": cid, "records": recs_sorted})

#     result.sort(key=lambda g: parse_date(g["records"][0]), reverse=True)

#     return result


# async def fetch_current_data(query) -> list[dict]:
#     """DB se current active (status=1, not deleted, aaj ke) records nikaal kar
#     JSON-safe dict list return karta hai."""
#     rows = await database2.fetch_all(query)
#     data = []
#     for row in rows:
#         d = dict(row)
#         for k, v in d.items():
#             if isinstance(v, datetime):
#                 d[k] = v.isoformat()
#         data.append(d)
#     return data


# @app.websocket("/ws/denial_alerts")
# async def websocket_escalation_alerts(websocket: WebSocket):
#     """Har baar poori current list (status=1, not deleted, aaj ke) bhejta hai,
#     latest added_date sabse upar. Status 2 hote hi record apne aap
#     list se hat jata hai kyunki wo query mein aata hi nahi.
#     DB mein koi bhi change hote hi (LISTEN/NOTIFY) turant push hoga.
#     Existing database2 pool se hi connection liya ja raha hai."""

#     await websocket.accept()

#     last_sent_ids: set = set()
#     notify_event = asyncio.Event()
#     db_connection = None   # databases.Connection wrapper
#     raw_conn = None        # underlying asyncpg connection (LISTEN ke liye)

#     query = """
#         SELECT *
#         FROM denial_escalation_master
#         WHERE escalate_status = '1' 
#         AND (is_deleted = FALSE OR is_deleted IS NULL)
#         AND added_date::date = CURRENT_DATE
#         ORDER BY added_date DESC;
#     """

#     async def send_if_changed():
#         """Latest data nikaal kar bhejta hai, sirf agar kuch change hua ho."""
#         nonlocal last_sent_ids
#         current_data = await fetch_current_data(query)
#         current_ids = {d["id"] for d in current_data}

#         if current_ids != last_sent_ids:
#             grouped_current = group_by_call_id(current_data)
#             await websocket.send_json({
#                 "type": "INITIAL_LOAD",
#                 "data": grouped_current
#             })
#             last_sent_ids = current_ids

#     def on_notify(conn, pid, channel, payload):
#         # DB se turant signal aata hai idhar (INSERT/UPDATE/DELETE hote hi)
#         notify_event.set()

#     try:
#         # ---------------------------------------------------------
#         # STEP 1: Initial load
#         # ---------------------------------------------------------
#         initial_data = await fetch_current_data(query)
#         last_sent_ids = {d["id"] for d in initial_data}

#         grouped_initial = group_by_call_id(initial_data)
#         await websocket.send_json({
#             "type": "INITIAL_LOAD",
#             "data": grouped_initial
#         })

#         # ---------------------------------------------------------
#         # STEP 2: database2 ke pool se hi ek connection nikaalo
#         # aur uska raw asyncpg connection le kar LISTEN lagao
#         # ---------------------------------------------------------
#         db_connection = database2.connection()
#         await db_connection.__aenter__()          # connection open karo (pool se)
#         raw_conn = db_connection.raw_connection    # underlying asyncpg.Connection
#         await raw_conn.add_listener("denial_escalation_channel", on_notify)

#         # ---------------------------------------------------------
#         # STEP 3: Jab bhi DB mein change ho (ya safety-net timeout),
#         # turant fresh data check karke bhejo
#         # ---------------------------------------------------------
#         while True:
#             try:
#                 # Notify ka wait — DB change hote hi ye turant trigger hoga
#                 await asyncio.wait_for(notify_event.wait(), timeout=30)
#             except asyncio.TimeoutError:
#                 pass  # 30 sec safety-net poll, agar kabhi notify miss ho jaye

#             notify_event.clear()
#             await send_if_changed()

#     except WebSocketDisconnect:
#         print("Frontend disconnected from WebSocket")
#     except Exception as e:
#         print(f"WebSocket Error: {e}")
#         import traceback
#         traceback.print_exc()
#         await websocket.close()
#     finally:
#         if raw_conn:
#             try:
#                 await raw_conn.remove_listener("denial_escalation_channel", on_notify)
#             except Exception:
#                 pass
#         if db_connection:
#             await db_connection.__aexit__(None, None, None)   # connection pool ko wapas

def group_by_call_id(records: list[dict]) -> list[dict]:
    """Same call_id wale records ko ek hi group/array mein daal deta hai,
    aur added_date ke hisaab se sabse latest sabse upar rakhta hai."""

    def parse_date(rec):
        val = rec.get("added_date")
        if isinstance(val, str):
            try:
                return datetime.fromisoformat(val)
            except (ValueError, TypeError):
                return datetime.min
        elif isinstance(val, datetime):
            return val
        return datetime.min

    grouped = {}
    for r in records:
        cid = r.get("call_id")
        grouped.setdefault(cid, []).append(r)

    result = []
    for cid, recs in grouped.items():
        recs_sorted = sorted(recs, key=parse_date, reverse=True)
        result.append({"call_id": cid, "records": recs_sorted})

    result.sort(key=lambda g: parse_date(g["records"][0]), reverse=True)

    return result


async def fetch_current_data(query) -> list[dict]:
    """DB se current active (status=1, not deleted) records nikaal kar
    JSON-safe dict list return karta hai."""
    rows = await database2.fetch_all(query)
    data = []
    for row in rows:
        d = dict(row)
        for k, v in d.items():
            if isinstance(v, datetime):
                d[k] = v.isoformat()
        data.append(d)
    return data


async def get_alert_flow_status(call_id) -> str | None:
    """alert_escalation_flow se status laata hai (fallback)."""
    if not call_id:
        return None
    q = """
        SELECT status
        FROM alert_escalation_flow
        WHERE call_id = :call_id
        LIMIT 1;
    """
    row = await database2.fetch_one(q, {"call_id": call_id})
    if row:
        return row.get("status")
    return None


async def attach_level_info_to_groups(grouped_data: list[dict]) -> list[dict]:
    """Har call_id group ke liye escalation level/role/severity attach karta hai."""
    for group in grouped_data:
        call_id = group.get("call_id")
        records = group.get("records", [])
        latest_record = records[0] if records else {}

        alert_flow_status = latest_record.get("alert_flow_status")
        if alert_flow_status is None:
            alert_flow_status = await get_alert_flow_status(call_id)

        group["alert_flow_status"] = alert_flow_status

        if await is_closed(call_id):
            action_details = await get_closed_details(call_id)
            current_level = to_int(action_details.get("action_by_level")) or 1
            group["is_closed"] = True
            group["action_details"] = action_details
        else:
            current_level = await get_or_init_escalation_level(
                call_id, latest_record.get("added_date"), is_denial=True
            )
            group["is_closed"] = False

        level_info = get_level_info_for_level(current_level)
        group["current_level"] = current_level
        group["current_role"] = level_info["role"]
        group["current_level_minutes"] = level_info["minutes"]
        group["severity"] = SEVERITY_BY_LEVEL.get(current_level, "LOW")

    return grouped_data


@app.websocket("/ws/denial_alerts")
async def websocket_denial_alerts(websocket: WebSocket):
    """URL se ?date=YYYY-MM-DD pass kar sakte hain.
    Date na diya to default CURRENT_DATE (aaj) use hoga."""

    await websocket.accept()

    last_sent_ids: set = set()
    notify_event = asyncio.Event()

    # --------------------------------------------------------
    # ✅ NAYA: URL se date param nikaalo
    # --------------------------------------------------------
    date_param = websocket.query_params.get("date")  # ?date=2025-01-15

    if date_param:
        try:
            selected_date = datetime.strptime(date_param, "%Y-%m-%d").date()
            date_condition = f"d.added_date::date = '{selected_date}'"
        except ValueError:
            await websocket.send_json({
                "type": "ERROR",
                "message": "Invalid date format! Use ?date=YYYY-MM-DD"
            })
            await websocket.close()
            return
    else:
        selected_date = None
        date_condition = "d.added_date::date = CURRENT_DATE"

    # --------------------------------------------------------
    # ✅ FIXED: Query me dynamic date condition
    # --------------------------------------------------------
    query = f"""
        SELECT d.*,
               af.status AS alert_flow_status
        FROM denial_escalation_master d
        LEFT JOIN LATERAL (
            SELECT status
            FROM alert_escalation_flow
            WHERE call_id = d.call_id
            LIMIT 1
        ) af ON TRUE
        WHERE d.escalate_status = '1' 
        AND (d.is_deleted = FALSE OR d.is_deleted IS NULL)
        AND {date_condition}
        ORDER BY d.added_date DESC;
    """

    async def send_if_changed():
        nonlocal last_sent_ids
        current_data = await fetch_current_data(query)
        current_ids = {d["id"] for d in current_data}

        if current_ids != last_sent_ids:
            grouped_current = group_by_call_id(current_data)
            grouped_current = await attach_level_info_to_groups(grouped_current)
            await websocket.send_json({
                "type": "INITIAL_LOAD",
                "count": len(grouped_current),
                "data": grouped_current
            })
            last_sent_ids = current_ids

    def on_notify(conn, pid, channel, payload):
        notify_event.set()

    try:
        # STEP 1: Initial load
        initial_data = await fetch_current_data(query)
        last_sent_ids = {d["id"] for d in initial_data}

        grouped_initial = group_by_call_id(initial_data)
        grouped_initial = await attach_level_info_to_groups(grouped_initial)

        await websocket.send_json({
            "type": "INITIAL_LOAD",
            "count": len(grouped_initial),
            "selected_date": str(selected_date) if selected_date else "today",
            "data": grouped_initial
        })

        # STEP 2: database2 connection & LISTEN
        async with database2.connection() as db_connection:
            raw_conn = db_connection.raw_connection
            await raw_conn.add_listener("denial_escalation_channel", on_notify)

            # STEP 3: Wait for notify and send changes
            while True:
                try:
                    await asyncio.wait_for(notify_event.wait(), timeout=30)
                except asyncio.TimeoutError:
                    pass

                notify_event.clear()
                await send_if_changed()

    except WebSocketDisconnect:
        print("Frontend disconnected from WebSocket")

    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"WebSocket Error: {e}\n{error_trace}")

        try:
            await websocket.send_json({
                "type": "ERROR",
                "message": str(e),
                "trace": error_trace
            })
        except:
            pass

        try:
            await websocket.close()
        except:
            pass

##########################################################################################
# ------------------------------------------------------------------
# Config
# ------------------------------------------------------------------
LEVEL_INFO_CENTRAL = {
    1: {"role": "MDT",                     "minutes": 0},
    2: {"role": "District Manager (DM)",   "minutes": 10},
    3: {"role": "Zonal Manager (ZM)",      "minutes": 20},
    4: {"role": "Operations Manager (OM)", "minutes": 30},
    5: {"role": "State Head (SH)",         "minutes": 45},
    6: {"role": "COO",                     "minutes": 60},
    7: {"role": "CBO",                     "minutes": 75},
}
SEVERITY_BY_LEVEL = {
    1: "LOW",       # MDT
    2: "LOW",       # DM
    3: "MEDIUM",    # ZM
    4: "MEDIUM",    # OM
    5: "HIGH",      # SH
    6: "HIGH",      # COO
    7: "CRITICAL",  # CBO
}
SORTED_LEVELS = sorted(
    LEVEL_INFO_CENTRAL.keys(),
    key=lambda l: LEVEL_INFO_CENTRAL[l]["minutes"]
)

# Redis keys
ESC_LEVEL_REDIS_PREFIX      = "esc_level:"
ESC_ESCALATED_AT_PREFIX     = "esc_escalated_at:"
ESC_ACTION_TAKEN_PREFIX     = "esc_action_taken:"
ESC_CLOSED_PREFIX           = "esc_closed:"
ESC_BUMP_LOCK_KEY           = "esc_bump_leader_lock"
ESC_BUMP_LOCK_TTL_MS        = 30000
_esc_bump_instance_id       = str(uuid.uuid4())

def normalize_vehicle_number(v) -> str:
    """
    Vehicle number ko ek standard format me laata hai.
    'TT 00 MP 0001' → 'TT00MP0001'
    'TT-00-MP-0001' → 'TT00MP0001'
    'tt00mp0001'    → 'TT00MP0001'
    
    Saare spaces, hyphens, dots, underscores hata ke uppercase kar deta hai.
    """
    if not v:
        return ""
    # String me convert, then saari special chars hata do, uppercase karo
    s = str(v).upper()
    for ch in [" ", "-", ".", "_", "/"]:
        s = s.replace(ch, "")
    return s.strip()


def extract_ambulance_number_from_payload(payload: dict) -> str:
    """
    Kisi bhi escalation payload se ambulance number nikalta hai
    aur normalize karke return karta hai.
    """
    # Direct fields
    amb = (
        payload.get("ambulance_no")
        or payload.get("amb_no")
        or payload.get("vehicleNumber")
    )

    # denial_record ke andar
    if not amb:
        dr = payload.get("denial_record") or {}
        amb = dr.get("amb_no") or dr.get("ambulance_no")

    # central_alerts list ke andar
    if not amb:
        for ca in (payload.get("central_alerts") or []):
            amb = ca.get("ambulance_no") or ca.get("amb_no")
            if amb:
                break

    # 👇 Normalize karke return karo
    return normalize_vehicle_number(amb)


import httpx

# FCM_SERVER_KEY = "AAAAdHbcA2w:APA91bGpaFIHWqD35tEQR0suCf_IRdOysTOvMsObjFgeIzGS_G2daBJjmRJrNyzQ13R5wrqBI9iVUTm-Ns_pIcs2R__m1s48RBNl__1FkFoQWAyUMZ4OPsDHNFg0a_rd2F9lXhHfInQB"   # 👈 yahan apni FCM key daalo

async def send_fcm_push(token: str, title: str, body: str, data: dict = None):
    """Firebase Cloud Messaging — SIRF data field bhejta hai.
    Android OS isko background me drop nahi karega, onMessageReceived call hoga."""
    if not token:
        logger.warning("FCM push skipped: no token provided")
        return

    try:
        firebase_admin.get_app()
    except ValueError:
        logger.error("❌ FCM ERROR: Firebase Admin SDK not initialized. Check service account JSON file.")
        return

    try:
        # 👇 SIRF data field — NO notification field
        clean_data = {
            "type": "Late",
            "title": title,
            "message": body,
            "discription": body,
        }

        if data:
            for k, v in data.items():
                clean_data[k] = str(v) if v is not None else ""

        # 👇 No notification object inside AndroidConfig, just high priority
        message = messaging.Message(
            token=token,
            data=clean_data,
            android=messaging.AndroidConfig(
                priority="high"
            ),
            apns=messaging.APNSConfig(
                headers={"apns-priority": "10"},
                payload=messaging.APNSPayload(
                    aps=messaging.Aps(
                        content_available=True
                    )
                )
            )
        )

        loop = asyncio.get_event_loop()
        response = await loop.run_in_executor(
            None,
            lambda: messaging.send(message)
        )

        logger.info(f"✅ FCM SUCCESS (Data-only): token={token[:20]}... msg_id={response} type=Late")

    except Exception as e:
        error_str = str(e)

        if "NOT_REGISTERED" in error_str or "invalid-registration-token" in error_str:
            logger.error(f"❌ FCM ERROR: Token invalid/expired: token={token[:20]}... error={error_str}")
        elif "permission-denied" in error_str:
            logger.error(f"❌ FCM ERROR: Permission denied: {error_str}")
        else:
            logger.exception(f"❌ send_fcm_push FAILED: {e}")

# ===========================================================================
# HELPERS (time → level)
# ===========================================================================
def get_level_for_elapsed_minutes(elapsed_minutes: int) -> int:
    current_level = 1
    for lvl in SORTED_LEVELS:
        if elapsed_minutes >= LEVEL_INFO_CENTRAL[lvl]["minutes"]:
            current_level = lvl
    return current_level


def get_level_info_for_level(level: int) -> dict:
    return LEVEL_INFO_CENTRAL.get(level, LEVEL_INFO_CENTRAL[1])


def elapsed_minutes_since(added_date) -> int:
    """Uses existing to_datetime() helper."""
    if isinstance(added_date, str):
        # 'T' ko space se replace karo aur microseconds (.) ke baad ka part hata do
        added_date = added_date.replace("T", " ").split(".")[0]
    
    dt = to_datetime(added_date)
    if not dt:
        return 0
    
    # 👇 Timezone fix: Agar datetime par timezone nahi hai, to use IST do
    if dt.tzinfo is None:
        dt = ist.localize(dt)
        
    # 👇 Ab current time aur DB time dono timezone-aware hain
    now = datetime.now(ist)
    
    return max(0, int((now - dt).total_seconds() // 60))


# ===========================================================================
# FETCHES (use existing serialize_row)
# ===========================================================================
async def fetch_central_alerts(call_id) -> list:
    rows = await database2.fetch_all(
        """
        SELECT alert_id, alert_type, incident_id, system_type, severity,
               ambulance_no, escalated_deny_remark, division, district,
               inc_latitude, inc_longitude, amb_lat, amb_long,
               inc_datetime, pilot_name, pilot_mobile,
               paramedic_name, paramedic_mobile, created_date,
               updated_date, escalate_status, is_deleted,
               escalated_date, cancel_date, escalated_by,
               cancel_by, remark
        FROM public.central_alerts
        WHERE CAST(alert_id AS text) = CAST(:call_id AS text)
        """,
        {"call_id": str(call_id)},   # 👈 yahan explicitly string banaya
    )
    return [serialize_row(r) for r in rows]


async def fetch_denial_record(call_id: str) -> dict:
    row = await database2.fetch_one(
        """
        SELECT id, mysql_id, call_id, amb_no, amb_default_mobile, caller_no,
               hp_name, challenge_val, meaning, denial_remark,
               alert_type, added_by, added_date
        FROM public.denial_escalation_master
        WHERE CAST(call_id AS text) = :call_id
        ORDER BY added_date DESC LIMIT 1   -- 👈 Yahan fix kiya
        """,
        {"call_id": str(call_id)},
    )
    return serialize_row(row) if row else {}

# ===========================================================================
# PAYLOAD BUILDER — single call_id ke liye combined payload (dono tables)
# ===========================================================================
async def build_escalation_payload(
    call_id: str,
    denial_record: dict,
    current_level: int,
    event_type: str = "new_escalation_alert",
    extra: dict = None,
    central_alerts_data: list = None  # 👈 Yeh add kiya
) -> dict:
    # 👇 Agar data pass hua to wahi use karo, warna DB se fetch karo
    if central_alerts_data is not None:
        central_alerts = central_alerts_data
    else:
        central_alerts = await fetch_central_alerts(call_id)
        
    level_info = get_level_info_for_level(current_level)
    severity = SEVERITY_BY_LEVEL.get(current_level, "LOW")

    payload = {
        "type": event_type,
        "call_id": call_id,
        "denial_record": denial_record,
        "central_alerts": central_alerts,
        "current_level": current_level,
        "current_role": level_info["role"],
        "current_level_minutes": level_info["minutes"],
        "severity": severity,
        "escalated_at": datetime.utcnow().isoformat(),
    }
    if extra:
        payload.update(extra)
    return payload

# ===========================================================================
async def attach_sla_info_to_alerts(alerts: list) -> list:
    """Each alert me SLA breach details add karta hai.
    
    Fetches from:
    - alert_thresholds table: SLA threshold for each alert_type + amb_area (rural/urban)
    - rtm_dashboard table: actual duration taken by ambulance
    
    Adds fields:
    - sla_threshold_seconds: SLA limit in seconds
    - sla_actual_seconds: actual time taken in seconds
    - sla_breach_seconds: how much time was exceeded
    - sla_breach_reason: human-readable reason
    - amb_area: "1" (rural) or "2" (urban)
    - amb_area_type: "Rural" or "Urban"
    """
    if not alerts:
        return alerts

    # 1. Saare thresholds fetch karo (cached, 30s TTL)
    try:
        thresholds = await get_active_thresholds()
    except Exception as e:
        logger.error(f"attach_sla: thresholds fetch failed: {e}")
        thresholds = []

    # 2. Saare unique incident_ids collect karo
    incident_ids = []
    for a in alerts:
        inc_id = str(a.get("incident_id") or a.get("call_id") or "")
        if inc_id and inc_id != "0" and inc_id not in incident_ids:
            incident_ids.append(inc_id)

    # 3. rtm_dashboard se data fetch karo for these incident_ids
    rtm_data = {}
    if incident_ids:
        try:
            placeholders = ", ".join([f":id{i}" for i in range(len(incident_ids))])
            params = {f"id{i}": incident_ids[i] for i in range(len(incident_ids))}
            query = f"""
                SELECT * FROM rtm_dashboard
                WHERE CAST(inc_ref_id AS text) IN ({placeholders})
            """
            rtm_rows = await database2.fetch_all(query, params)
            for r in rtm_rows:
                rtm_data[str(r["inc_ref_id"])] = normalize_row(r)
        except Exception as e:
            logger.error(f"attach_sla: rtm_dashboard fetch failed: {e}")

    # 4. Alert type → rtm_dashboard duration field mapping
    ALERT_DURATION_MAP = {
        "ACK_DELAY": "acknowledge_duration",
        "START_DELAY": "start_from_base_duration",
        "AT_SCENE_DELAY": "at_scene_duration",
    }

    # 5. Each alert me SLA info add karo
    for alert in alerts:
        alert_type = alert.get("alert_type", "")
        incident_id = str(alert.get("incident_id") or alert.get("call_id") or "")

        # rtm_dashboard se row uthao
        rtm_row = rtm_data.get(incident_id, {})

        # amb_area (1=rural, 2=urban)
        amb_area = rtm_row.get("amb_working_area")
        amb_area_type = ""
        if str(amb_area) == "1":
            amb_area_type = "Rural"
        elif str(amb_area) == "2":
            amb_area_type = "Urban"

        # Threshold dhoondho — alert_type + amb_area match
        # null amb_area means applies to ALL areas
        threshold_seconds = None
        for t in thresholds:
            if t.get("alert_type") != alert_type:
                continue
            t_amb_area = t.get("amb_area")
            if t_amb_area is not None and str(t_amb_area) != str(amb_area):
                continue
            threshold_seconds = int(t.get("threshold_seconds") or 0)
            break

        # Actual duration calculate karo
        actual_seconds = None
        breach_reason = ""

        # =========================================================
        # MDT_NOT_LOGGED_IN
        # =========================================================
        if alert_type == "MDT_NOT_LOGGED_IN":
            pilot_login_out = rtm_row.get("pilot_login_out")
            actual_seconds = 0
            breach_reason = (
                f"MDT not logged in — pilot login status: {pilot_login_out}"
            )

        # =========================================================
        # BACK_TO_BASE_DELAY — patient_handover ke baad se check
        # =========================================================
        elif alert_type == "BACK_TO_BASE_DELAY":
            patient_handover_dt = to_datetime(rtm_row.get("patient_handover"))
            back_to_base_dt = to_datetime(rtm_row.get("back_to_base_loc"))

            if patient_handover_dt and back_to_base_dt:
                diff = (back_to_base_dt - patient_handover_dt).total_seconds()
                if diff >= 0:
                    actual_seconds = int(diff)

            if actual_seconds is not None and threshold_seconds is not None:
                breach_seconds = max(0, actual_seconds - threshold_seconds)
                breach_reason = (
                    f"Back to base delay — ambulance took {format_seconds_human(actual_seconds)} "
                    f"to return to base after patient handover, "
                    f"{format_seconds_human(breach_seconds)} over the "
                    f"{format_seconds_human(threshold_seconds)} SLA limit"
                    + (f" ({amb_area_type} area)" if amb_area_type else "")
                )
            elif actual_seconds is not None and threshold_seconds is None:
                breach_reason = (
                    f"Back to base delay — ambulance took {format_seconds_human(actual_seconds)} "
                    f"to return to base after patient handover. "
                    f"SLA threshold not configured for this alert type."
                )
            else:
                breach_reason = (
                    "Back to base delay — patient handover or back to base "
                    "time data not available in rtm_dashboard."
                )

        # =========================================================
        # ACK_DELAY, START_DELAY, AT_SCENE_DELAY — direct duration field
        # =========================================================
        else:
            duration_field = ALERT_DURATION_MAP.get(alert_type)
            if duration_field:
                raw_duration = rtm_row.get(duration_field)
                actual_seconds = hhmmss_to_seconds(raw_duration) if raw_duration else None

                if actual_seconds is not None and threshold_seconds is not None:
                    breach_seconds = max(0, actual_seconds - threshold_seconds)

                    # Human-readable message per alert type
                    if alert_type == "ACK_DELAY":
                        breach_reason = (
                            f"Acknowledge delay — ambulance took {format_seconds_human(actual_seconds)} "
                            f"to acknowledge the call, "
                            f"{format_seconds_human(breach_seconds)} over the "
                            f"{format_seconds_human(threshold_seconds)} SLA limit"
                            + (f" ({amb_area_type} area)" if amb_area_type else "")
                        )
                    elif alert_type == "START_DELAY":
                        breach_reason = (
                            f"Start from base delay — ambulance took {format_seconds_human(actual_seconds)} "
                            f"to depart from base, "
                            f"{format_seconds_human(breach_seconds)} over the "
                            f"{format_seconds_human(threshold_seconds)} SLA limit"
                            + (f" ({amb_area_type} area)" if amb_area_type else "")
                        )
                    elif alert_type == "AT_SCENE_DELAY":
                        breach_reason = (
                            f"At scene delay — ambulance spent {format_seconds_human(actual_seconds)} "
                            f"at the scene, "
                            f"{format_seconds_human(breach_seconds)} over the "
                            f"{format_seconds_human(threshold_seconds)} SLA limit"
                            + (f" ({amb_area_type} area)" if amb_area_type else "")
                        )
                    else:
                        breach_reason = (
                            f"{alert_type} — actual: {format_seconds_human(actual_seconds)}, "
                            f"threshold: {format_seconds_human(threshold_seconds)}, "
                            f"breached by: {format_seconds_human(breach_seconds)}"
                            + (f" ({amb_area_type} area)" if amb_area_type else "")
                        )
                elif actual_seconds is not None and threshold_seconds is None:
                    breach_reason = (
                        f"{alert_type} — ambulance took {format_seconds_human(actual_seconds)}. "
                        f"SLA threshold not configured for this alert type."
                    )
                else:
                    breach_reason = (
                        f"{alert_type} — duration data not available in rtm_dashboard."
                    )
            else:
                breach_reason = f"{alert_type} — unknown alert type, no SLA mapping available."

        # breach seconds
        breach_seconds = None
        if actual_seconds is not None and threshold_seconds is not None:
            breach_seconds = max(0, actual_seconds - threshold_seconds)

        # 👇 SLA fields add karo alert me
        alert["sla_threshold_seconds"] = threshold_seconds
        alert["sla_actual_seconds"] = actual_seconds
        alert["sla_breach_seconds"] = breach_seconds
        alert["sla_breach_reason"] = breach_reason
        alert["amb_area"] = amb_area
        alert["amb_area_type"] = amb_area_type

    return alerts

# ===========================================================================
# PAYLOAD BUILDER
# ===========================================================================
async def build_all_escalation_payloads(requested_level: int = None) -> list:
    """Saare alerts (central + denial) ko ek hi list mein laata hai,
    unhe time ke hisaab se DESC (latest upar) sort karta hai."""

    # Fetch denial records (sirf un case mein lao jab level 1 na ho)
    denial_rows = []
    if requested_level is None or requested_level != 1:
        # 👇 FIX: DISTINCT ON use kiya taaki ek call_id ka sirf 1 latest record aaye
        denial_rows = await database2.fetch_all(
            """
            SELECT * FROM (
                SELECT DISTINCT ON (call_id) 
                       id, mysql_id, call_id, amb_no, amb_default_mobile, caller_no,
                       hp_name, challenge_val, meaning, denial_remark,
                       alert_type, added_by, added_date
                FROM public.denial_escalation_master
                WHERE added_date::date = CURRENT_DATE
                ORDER BY call_id, added_date DESC
            ) as latest_denials
            ORDER BY added_date DESC
            LIMIT 5000
            """
        )

    # Fetch central alerts (hamesha laao)
    central_rows = await database2.fetch_all(
        """
        SELECT *
        FROM public.central_alerts
        WHERE created_date::date = CURRENT_DATE
        AND escalate_status = '1'
        AND is_deleted = false
        ORDER BY created_date DESC
        LIMIT 5000
        """
    )

    unified_alerts = []

    # 1. Denial records ko unified list mein daalo
    for record in denial_rows:
        d = serialize_row(record)
        d["record_source"] = "denial"
        d["sort_time"] = d.get("added_date")
        unified_alerts.append(d)

    # 2. Central alerts ko unified list mein daalo
    for record in central_rows:
        c = serialize_row(record)
        c["record_source"] = "central"
        c["sort_time"] = c.get("created_date")
        unified_alerts.append(c)

    # 3. Saare alerts ko time ke hisaab se DESC (latest upar) sort karo
    unified_alerts.sort(key=lambda x: x.get("sort_time") or "", reverse=True)

    logger.info(
        f"BUILD PAYLOADS: fetched denial={len(denial_rows)}, "
        f"central={len(central_rows)}, requested_level={requested_level}"
    )

    # 4. Level info attach karo aur filter lagao
    filtered_alerts = []
    for alert in unified_alerts:
        # Central alert ke liye alert_id use karo, Denial ke liye call_id
        if alert.get("record_source") == "central":
            call_id = str(alert.get("alert_id") or "")
        else:
            call_id = str(alert.get("call_id") or "")

        if not call_id or call_id == "0":
            continue

        if await is_closed(call_id):
            action_details = await get_closed_details(call_id)
            current_level = to_int(action_details.get("action_by_level")) or 1
            alert["is_closed"] = True
            alert["action_details"] = action_details

            if requested_level is not None and current_level != requested_level:
                continue
        else:
            is_denial = (alert.get("record_source") == "denial")
            current_level = await get_or_init_escalation_level(
                call_id,
                alert.get("sort_time"),
                is_denial=is_denial
            )
            alert["is_closed"] = False

            if requested_level is not None and current_level != requested_level:
                continue

        level_info = get_level_info_for_level(current_level)
        alert["current_level"] = current_level
        alert["current_role"] = level_info["role"]
        alert["current_level_minutes"] = level_info["minutes"]
        alert["severity"] = SEVERITY_BY_LEVEL.get(current_level, "LOW")
        alert["type"] = "current_escalation"
        alert["escalated_at"] = datetime.now(ist).isoformat()

        # 👇 FIX: Level 1 ke liye faltu fields hata do
        if requested_level == 1:
            alert.pop("escalated_at", None)
            alert.pop("updated_date", None)
            alert.pop("current_level_minutes", None)

        filtered_alerts.append(alert)

    # SLA breach details add karo
    filtered_alerts = await attach_sla_info_to_alerts(filtered_alerts)

    logger.info(
        f"BUILD PAYLOADS: returning {len(filtered_alerts)} alerts "
        f"for level={requested_level}"
    )

    return filtered_alerts
# ===========================================================================
# REDIS STATE — per call_id current level + closed flag
# ===========================================================================
async def get_or_init_escalation_level(call_id: str, added_date, is_denial: bool = False) -> int:
    key = f"{ESC_LEVEL_REDIS_PREFIX}{call_id}"
    raw = await redis_client.get(key)

    current_level = None
    if raw is not None:
        val = raw if isinstance(raw, str) else raw.decode()
        current_level = to_int(val)

    # Denial records hamesha level 2 se start hote hain
    if is_denial and (current_level is None or current_level < 2):
        current_level = 2
        await redis_client.set(key, str(current_level))
        await redis_client.set(
            f"{ESC_ESCALATED_AT_PREFIX}{call_id}",
            datetime.utcnow().isoformat()
        )
        return current_level

    # 👇 HAMESHA time-based level calculate karo (sticky hone ka fix)
    elapsed = elapsed_minutes_since(added_date)
    time_based_level = get_level_for_elapsed_minutes(elapsed)
    if is_denial and time_based_level < 2:
        time_based_level = 2
    # Cap at 7
    time_based_level = min(time_based_level, 7)

    # 👇 Agar Redis me level hai hi nahi → time-based use karo + set karo
    if current_level is None:
        current_level = time_based_level
        await redis_client.set(key, str(current_level))
        await redis_client.set(
            f"{ESC_ESCALATED_AT_PREFIX}{call_id}",
            datetime.utcnow().isoformat()
        )
        return current_level

    # 👇 Agar Redis me level hai but time-based level ZYADA hai → upgrade kar do
    #    (Lekin kabhi downgrade mat karo — agar manually level 7 set kiya hai to rehne do)
    if time_based_level > current_level:
        current_level = time_based_level
        await redis_client.set(key, str(current_level))
        await redis_client.set(
            f"{ESC_ESCALATED_AT_PREFIX}{call_id}",
            datetime.utcnow().isoformat()
        )
        logger.info(
            f"LEVEL RECALC: call_id={call_id} upgraded to {current_level} "
            f"({LEVEL_INFO_CENTRAL[current_level]['role']}, elapsed={elapsed} min)"
        )

    return current_level

async def set_escalation_level(call_id: str, level: int):
    await redis_client.set(f"{ESC_LEVEL_REDIS_PREFIX}{call_id}", str(level))
    await redis_client.set(
        f"{ESC_ESCALATED_AT_PREFIX}{call_id}",
        datetime.utcnow().isoformat()
    )


async def is_action_taken(call_id: str) -> bool:
    return (await redis_client.get(f"{ESC_ACTION_TAKEN_PREFIX}{call_id}")) is not None


async def is_closed(call_id: str) -> bool:
    return (await redis_client.get(f"{ESC_CLOSED_PREFIX}{call_id}")) is not None


async def get_closed_details(call_id: str) -> dict:
    raw = await redis_client.get(f"{ESC_CLOSED_PREFIX}{call_id}")
    if not raw:
        return {}
    try:
        val = raw if isinstance(raw, str) else raw.decode()
        return json.loads(val)
    except Exception:
        return {}


async def mark_action_taken(
    call_id: str,
    action_by: str,
    action_by_role: str,
    action_by_level: int,
    action_remark: str,
    action_type: str = "acknowledge",
) -> dict:
    """
    Kisi role ne action liya:
    1. esc_action_taken:{call_id} set karo
    2. esc_closed:{call_id}      set karo
    3. esc_level:* & esc_escalated_at:* delete karo → aage escalate nahi hoga
    """
    action_details = {
        "call_id": call_id,
        "action_by": action_by,
        "action_by_role": action_by_role,
        "action_by_level": action_by_level,
        "action_remark": action_remark,
        "action_type": action_type,
        "action_taken_at": datetime.utcnow().isoformat(),
    }
    action_json = json.dumps(action_details)

    await redis_client.set(f"{ESC_ACTION_TAKEN_PREFIX}{call_id}", action_json)
    await redis_client.set(f"{ESC_CLOSED_PREFIX}{call_id}", action_json)

    # Cleanup level state — aage escalate NAHI karna
    await redis_client.delete(f"{ESC_LEVEL_REDIS_PREFIX}{call_id}")
    await redis_client.delete(f"{ESC_ESCALATED_AT_PREFIX}{call_id}")

    return action_details


# Escalation Flow Table Mapping
ESC_FLOW_LEVEL_MAP = {
    1: "pilot",
    2: "dm",
    3: "zm",
    4: "om",   
    5: "sh",   
    6: "coo",
    7: "cbo",
}

ROLE_NAMES = {
    1: "Pilot",
    2: "DM",
    3: "ZM",
    4: "OM",
    5: "SH",
    6: "COO",
    7: "CBO"
}

def safe_int(val):
    """Convert string/float to int safely, return None if invalid."""
    if val is None or str(val).strip() == "" or str(val).strip() == "None":
        return None
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return None

async def update_escalation_flow(
    call_id: str,
    alert_id: str,
    level: int,
    event_type: str,
    action_by: str = None,
    action_remark: str = None,
    generated_at: str = None
):
    """alert_escalation_flow table me real-time tracking maintain karta hai."""
    try:
        # 👇 Dono ko int me convert karo
        c_id = safe_int(call_id)
        a_id = safe_int(alert_id)
        lvl = safe_int(level)
        
        # Timestamp convert karo
        gen_dt = to_datetime(generated_at) if generated_at else None
        
        # Build WHERE clause (Sirf wahi ID search karenge jo real me available hai aur 0 nahi hai)
        where_clauses = []
        params = {}
        if a_id is not None and a_id != 0:
            where_clauses.append("alert_id = :a_id")
            params["a_id"] = a_id
        if c_id is not None and c_id != 0:
            where_clauses.append("call_id = :c_id")
            params["c_id"] = c_id
            
        if not where_clauses:
            return # Koi real ID nahi hai track karne ke liye

        query = f"""
            SELECT sr_no FROM alert_escalation_flow 
            WHERE {' OR '.join(where_clauses)}
            LIMIT 1
        """
        existing = await database2.fetch_one(query, params)

        if event_type == "NEW_ALERT":
            if not existing:
                role_prefix = ESC_FLOW_LEVEL_MAP.get(level)
                role_name = ROLE_NAMES.get(level, "Unknown")
                
                cols = ["alert_id", "call_id", "alert_generated_at", "level", "status", "is_closed", "created_at", "updated_at"]
                vals = [":alert_id", ":call_id", ":alert_generated_at", ":level", ":status", ":is_closed", "NOW()", "NOW()"]
                
                # 👇 Agar value None hai to Integer 0 daal do (DB NOT NULL constraint ke liye)
                params_insert = {
                    "alert_id": a_id if a_id is not None else 0,
                    "call_id": c_id if c_id is not None else 0,
                    "alert_generated_at": gen_dt,
                    "level": lvl,
                    "status": f"Escalated To {role_name}",
                    "is_closed": 0
                }

                if role_prefix:
                    cols.append(f"{role_prefix}_send_datetime")
                    vals.append("NOW()")

                sql = f"""
                    INSERT INTO alert_escalation_flow ({', '.join(cols)})
                    VALUES ({', '.join(vals)})
                """
                await database2.execute(sql, params_insert)

        elif event_type == "LEVEL_BUMP":
            if not existing:
                # Fallback if bumped before insert
                await update_escalation_flow(call_id, alert_id, level, "NEW_ALERT", generated_at=generated_at)
                existing = await database2.fetch_one(query, params)
                if not existing: 
                    return

            role_prefix = ESC_FLOW_LEVEL_MAP.get(level)
            role_name = ROLE_NAMES.get(level, "Unknown")
            params_update = {
                "sr_no": existing["sr_no"],
                "level": lvl,
                "status": f"Escalated To {role_name}"
            }
            if role_prefix:
                sql = f"""
                    UPDATE alert_escalation_flow 
                    SET level = :level, status = :status, {role_prefix}_send_datetime = NOW(), updated_at = NOW()
                    WHERE sr_no = :sr_no
                """
            else:
                sql = """
                    UPDATE alert_escalation_flow 
                    SET level = :level, status = :status, updated_at = NOW()
                    WHERE sr_no = :sr_no
                """
            await database2.execute(sql, params_update)

        elif event_type == "ACTION_TAKEN":
            if not existing:
                return

            role_prefix = ESC_FLOW_LEVEL_MAP.get(level)
            params_update = {
                "sr_no": existing["sr_no"],
                "action_by": action_by,
                "status": "Closed",
                "is_closed": 1,
                "action_remark": action_remark
            }
            if role_prefix:
                sql = f"""
                    UPDATE alert_escalation_flow 
                    SET status = :status, is_closed = :is_closed, action_by = :action_by, 
                        {role_prefix}_action = :action_remark, {role_prefix}_action_datetime = NOW(),
                        updated_at = NOW()
                    WHERE sr_no = :sr_no
                """
            else:
                sql = """
                    UPDATE alert_escalation_flow 
                    SET status = :status, is_closed = :is_closed, action_by = :action_by, updated_at = NOW()
                    WHERE sr_no = :sr_no
                """
            await database2.execute(sql, params_update)

    except Exception as e:
        logger.error(f"Failed to update escalation flow: {e}")
# ===========================================================================
# REDIS LEADER LOCK (bump watcher ke liye)
# ===========================================================================
async def try_acquire_bump_leadership() -> bool:
    if await redis_client.set(
        ESC_BUMP_LOCK_KEY, _esc_bump_instance_id,
        nx=True, px=ESC_BUMP_LOCK_TTL_MS
    ):
        return True
    current = await redis_client.get(ESC_BUMP_LOCK_KEY)
    if isinstance(current, bytes):
        current = current.decode()
    if current == _esc_bump_instance_id:
        await redis_client.pexpire(ESC_BUMP_LOCK_KEY, ESC_BUMP_LOCK_TTL_MS)
        return True
    return False


# ===========================================================================
# WEBSOCKET — /ws/escalation_alerts (path-based level filter: /1, /2, ..., /7)
# ===========================================================================
# ===========================================================================
# WEBSOCKET — /ws/escalation_alerts (single endpoint, level as query param)
#   - /ws/escalation_alerts              → saare levels ke alerts
#   - /ws/escalation_alerts?level=2      → sirf level 2 (DM) ke alerts
# ===========================================================================
@app.websocket("/ws/escalation_alerts")
async def websocket_escalation_alerts(
    websocket: WebSocket,
    level: Optional[int] = Query(default=None),   # None = all, 1-7 = specific
    user_id: str = "anonymous"
):
    """
    Single endpoint jo dono kaam karta hai:
      /ws/escalation_alerts              → saare levels (LOW/MEDIUM/HIGH/CRITICAL sab)
      /ws/escalation_alerts?level=2      → sirf level 2 (DM — LOW severity)
      /ws/escalation_alerts?level=7      → sirf level 7 (CBO — CRITICAL severity)
    """
    if level is not None and (level < 1 or level > 7):
        await websocket.accept()
        await websocket.send_json({
            "type": "ERROR",
            "message": f"Invalid level: {level}. Must be 1-7.",
            "valid_levels": {
                1: "MDT", 2: "DM", 3: "ZM", 4: "OM",
                5: "SH", 6: "COO", 7: "CBO"
            }
        })
        await websocket.close(code=1008)
        return

    await _handle_escalation_ws(websocket, requested_level=level, user_id=user_id)



# ---------------------------------------------------------------------------
# Common handler — dono routes yahan se chalte hain
# ---------------------------------------------------------------------------
async def _handle_escalation_ws(websocket: WebSocket, requested_level: int, user_id: str):
    conn_info = await manager.connect(websocket, user_id)
    queue = conn_info["queue"]

    vehicle_number = None
    vehicle_number_raw = None
    fcm_token = None

    # =========================================================
    # Level 1: SILENTLY wait for registration message
    # =========================================================
    if requested_level == 1:
        try:
            raw = await asyncio.wait_for(websocket.receive_text(), timeout=60)
            data = json.loads(raw)

            vehicle_number_raw = (
                data.get("vehicleNumber")
                or data.get("ambulance_no")
                or data.get("ambulanceNumber")
            )
            fcm_token = data.get("token") or data.get("fcm_token")

            vehicle_number = normalize_vehicle_number(vehicle_number_raw)

            if not vehicle_number or not fcm_token:
                await websocket.close(code=1008)
                return

            # Redis me token save karo
            await redis_client.set(f"fcm_token:{vehicle_number}", fcm_token)
            logger.info(
                f"Level-1 MDT registered & SAVED IN REDIS: user={user_id}, "
                f"raw={vehicle_number_raw}, normalized={vehicle_number}"
            )

        except asyncio.TimeoutError:
            await websocket.close(code=1008)
            return
        except WebSocketDisconnect:
            return
        except Exception as e:
            logger.exception(f"Level-1 registration failed: {e}")
            try:
                await websocket.close(code=1011)
            except Exception:
                pass
            return

    try:
        # Send INITIAL_LOAD (First time connect hote hi)
        await send_current_escalations(
            websocket,
            requested_level=requested_level,
            vehicle_filter=vehicle_number,
            fcm_token=None
        )

        async def drain_loop():
            """Sirf INITIAL_LOAD format bhejo — kuch aur nahi."""
            last_refresh_time = 0
            notified_alert_ids = set()
            last_sent_signature = None  # 👈 SPAM STOPPER

            while True:
                payload = await queue.get()
                msg_type = payload.get("type")

                # Level 1 ke liye ye events allow karo (ALL_ALERTS wapas laaya)
                allowed_types = [
                    "ESCALATION_REFRESH",
                    "ALL_ALERTS",
                    "new_escalation_alert",
                    "escalation_level_changed",
                    "escalation_closed",
                    "current_escalation",
                ]

                if msg_type not in allowed_types:
                    continue

                # Throttle (0.5s for Level 1 to be very fast)
                now = time.time()
                throttle_time = 0.5 if requested_level == 1 else 1
                if now - last_refresh_time < throttle_time:
                    continue
                last_refresh_time = now

                # Re-fetch and send INITIAL_LOAD (same format always)
                try:
                    payloads = await build_all_escalation_payloads(requested_level)

                    if requested_level == 1 and vehicle_number:
                        filtered = []
                        for p in payloads:
                            amb_raw = (
                                p.get("ambulance_no")
                                or p.get("amb_no")
                                or p.get("vehicleNumber")
                            )
                            amb_normalized = normalize_vehicle_number(amb_raw)
                            if amb_normalized == vehicle_number:
                                filtered.append(p)
                        payloads = filtered

                        # 👇 SPAM STOPPER LOGIC: Check if data actually changed for our vehicle
                        current_signature = set()
                        for p in payloads:
                            uid = str(p.get("alert_id") or p.get("call_id") or p.get("incident_id"))
                            lvl = p.get("current_level")
                            is_cl = p.get("is_closed")
                            # Signature banao: Alert ID + Level + Closed Status
                            current_signature.add(f"{uid}_{lvl}_{is_cl}")
                        
                        # Agar data same hai, to message bhejne ki zaroorat nahi (Spam Stop)
                        if current_signature == last_sent_signature:
                            continue
                        
                        # Data change ho gaya hai, last_sent_signature update karo
                        last_sent_signature = current_signature

                        # 👇 FCM PUSH LOGIC: Naye alerts ke liye
                        if fcm_token and len(payloads) > 0:
                            for p in payloads:
                                unique_id = str(
                                    p.get("alert_id") 
                                    or p.get("call_id") 
                                    or p.get("incident_id") 
                                    or ""
                                )
                                
                                if unique_id and unique_id not in notified_alert_ids:
                                    logger.info(f"🚀 Triggering FCM from WS loop for ID={unique_id}")
                                    await send_fcm_push(
                                        token=fcm_token,
                                        title=f"Ambulance {vehicle_number_raw} - New Alert",
                                        body=f"Alert: {p.get('alert_type', 'Escalation')}",
                                        data={
                                            "alert_id": unique_id,
                                            "type": "Late",
                                        }
                                    )
                                    notified_alert_ids.add(unique_id)

                    await manager.safe_send(websocket, {
                        "type": "INITIAL_LOAD",
                        "count": len(payloads),
                        "data": payloads,
                        "vehicle_filter": vehicle_number if requested_level == 1 else None,
                    })
                except Exception as e:
                    logger.error(f"drain_loop refresh failed: {e}")

        async def keep_alive_loop():
            while True:
                try:
                    await websocket.receive_text()
                except WebSocketDisconnect:
                    break
                except Exception:
                    break

        sender   = asyncio.create_task(drain_loop())
        receiver = asyncio.create_task(keep_alive_loop())

        done, pending = await asyncio.wait(
            [sender, receiver], return_when=asyncio.FIRST_COMPLETED
        )
        for t in pending:
            t.cancel()
        for t in done:
            if t.exception():
                logger.exception(f"WS task FAILED: {t.exception()}")

    except WebSocketDisconnect:
        pass
    except Exception as e:
        logger.exception(f"websocket_escalate_alerts FAILED: {e}")
    finally:
        manager.disconnect(websocket)

async def send_current_escalations(
    websocket: WebSocket,
    requested_level: int = None,
    vehicle_filter: str = None,
    fcm_token: str = None
):
    """Sirf INITIAL_LOAD format bhejta hai — kuch aur nahi."""
    try:
        payloads = await build_all_escalation_payloads(requested_level)

        if requested_level == 1:
            logger.info(
                f"LEVEL 1 SEND: build_all returned {len(payloads)} alerts, "
                f"vehicle_filter={vehicle_filter}"
            )

        if vehicle_filter:
            normalized_filter = normalize_vehicle_number(vehicle_filter)
            filtered = []
            for p in payloads:
                amb_raw = (
                    p.get("ambulance_no")
                    or p.get("amb_no")
                    or p.get("vehicleNumber")
                )
                amb_normalized = normalize_vehicle_number(amb_raw)

                logger.info(
                    f"  VEHICLE MATCH: payload_amb={amb_raw} → "
                    f"normalized={amb_normalized}, filter={normalized_filter}, "
                    f"match={amb_normalized == normalized_filter}"
                )

                if amb_normalized == normalized_filter:
                    filtered.append(p)
            payloads = filtered

        # 👇 ONLY this format — nothing else
        await websocket.send_json({
            "type": "INITIAL_LOAD",
            "count": len(payloads),
            "data": payloads,
            "vehicle_filter": vehicle_filter,
        })

        filter_name = (
            LEVEL_INFO_CENTRAL[requested_level]["role"]
            if requested_level else "ALL"
        )
        logger.info(
            f"Escalation WS: sent {len(payloads)} records in ONE message "
            f"(filter: {filter_name}, vehicle: {vehicle_filter})"
        )

    except Exception as e:
        # 👇 No ERROR message to client — just log
        logger.exception(f"send_current_escalations FAILED: {e}")
# ===========================================================================
# WATCHER: TIME-BASED LEVEL BUMP (MDT → DM → ZM → OM → SH → COO → CBO)
#   *** AGAR ACTION LIYA TO AAGE ESCALATE NAHI KAREGA ***
# ===========================================================================
async def init_escalation_levels_for_today():
    """Aaj ke saare alerts scan karke Redis me level set karo."""
    try:
        # ---------- Central Alerts ----------
        # 👇 FIX: alert_id uthao
        central_rows = await database2.fetch_all(
            """
            SELECT alert_id, created_date
            FROM public.central_alerts
            WHERE created_date::date = CURRENT_DATE
            AND escalate_status = '1'
            AND is_deleted = false
            """
        )

        central_init = 0
        for row in central_rows:
            call_id = str(row["alert_id"]) # 👈 FIX
            if not call_id or call_id == "0":
                continue

            key = f"{ESC_LEVEL_REDIS_PREFIX}{call_id}"
            exists = await redis_client.get(key)

            if exists is None:
                await get_or_init_escalation_level(
                    call_id,
                    row["created_date"],
                    is_denial=False
                )
                central_init += 1

        # ---------- Denial Records ----------
        denial_rows = await database2.fetch_all(
            """
            SELECT call_id, added_date
            FROM public.denial_escalation_master
            WHERE added_date::date = CURRENT_DATE
            """
        )

        denial_init = 0
        for row in denial_rows:
            call_id = str(row["call_id"])
            if not call_id or call_id == "0":
                continue

            key = f"{ESC_LEVEL_REDIS_PREFIX}{call_id}"
            exists = await redis_client.get(key)

            if exists is None:
                await get_or_init_escalation_level(
                    call_id,
                    row["added_date"],
                    is_denial=True
                )
                denial_init += 1

        if central_init > 0 or denial_init > 0:
            logger.info(
                f"INIT ESCALATION: initialized {central_init} central + "
                f"{denial_init} denial Redis keys (total {central_init + denial_init})"
            )

    except Exception as e:
        logger.error(f"init_escalation_levels_for_today failed: {e}")


async def escalation_level_bump_watcher():
    logger.info("Escalation Level Bump Watcher STARTED")

    cursor = 0
    last_init_time = 0

    while True:
        try:
            if not await try_acquire_bump_leadership():
                await asyncio.sleep(10)
                continue

            now = time.time()
            if now - last_init_time > 300:
                await init_escalation_levels_for_today()
                last_init_time = now

            cursor, keys = await redis_client.scan(
                cursor=cursor,
                match=f"{ESC_LEVEL_REDIS_PREFIX}*",
                count=100,
            )

            for key in keys:
                if isinstance(key, bytes):
                    key = key.decode()
                call_id = key.replace(ESC_LEVEL_REDIS_PREFIX, "")

                if await is_closed(call_id) or await is_action_taken(call_id):
                    await redis_client.delete(f"{ESC_LEVEL_REDIS_PREFIX}{call_id}")
                    await redis_client.delete(f"{ESC_ESCALATED_AT_PREFIX}{call_id}")
                    continue

                raw_level = await redis_client.get(key)
                val = raw_level if isinstance(raw_level, str) else (raw_level.decode() if raw_level else None)
                current_level = to_int(val)
                if current_level is None:
                    continue

                if current_level >= 7:
                    continue

                d = await fetch_denial_record(call_id)
                is_denial_rec = True

                if not d:
                    is_denial_rec = False
                    # 👇 FIX: alert_id se search karo
                    central_row = await database2.fetch_one(
                        """SELECT * FROM public.central_alerts 
                          WHERE CAST(alert_id AS text) = :id
                          ORDER BY created_date DESC LIMIT 1""",
                          {"id": str(call_id)} # 👈 FIX
                    )
                    if not central_row:
                        await redis_client.delete(f"{ESC_LEVEL_REDIS_PREFIX}{call_id}")
                        await redis_client.delete(f"{ESC_ESCALATED_AT_PREFIX}{call_id}")
                        logger.warning(f"STALE KEY CLEANUP: call_id={call_id} (no DB record found)")
                        continue
                    d = serialize_row(central_row)

                check_date = d.get("added_date") if is_denial_rec else d.get("created_date")
                if not check_date:
                    continue

                elapsed = elapsed_minutes_since(check_date)

                time_based_level = get_level_for_elapsed_minutes(elapsed)
                if is_denial_rec and time_based_level < 2:
                    time_based_level = 2
                time_based_level = min(time_based_level, 7)

                if time_based_level <= current_level:
                    continue

                new_level = time_based_level

                if await is_closed(call_id) or await is_action_taken(call_id):
                    await redis_client.delete(f"{ESC_LEVEL_REDIS_PREFIX}{call_id}")
                    await redis_client.delete(f"{ESC_ESCALATED_AT_PREFIX}{call_id}")
                    continue

                await set_escalation_level(call_id, new_level)

                if is_denial_rec:
                    payload = await build_escalation_payload(
                        call_id=call_id,
                        denial_record=d,
                        current_level=new_level,
                        event_type="escalation_level_changed",
                        extra={
                            "previous_level": current_level,
                            "previous_role":  LEVEL_INFO_CENTRAL[current_level]["role"],
                            "jumped_levels":  new_level - current_level,
                            "elapsed_minutes": elapsed,
                        },
                    )
                else:
                    payload = await build_escalation_payload(
                        call_id=call_id,
                        denial_record={},
                        central_alerts_data=[d],
                        current_level=new_level,
                        event_type="escalation_level_changed",
                        extra={
                            "previous_level": current_level,
                            "previous_role":  LEVEL_INFO_CENTRAL[current_level]["role"],
                            "jumped_levels":  new_level - current_level,
                            "elapsed_minutes": elapsed,
                        },
                    )

                manager.broadcast(payload)

                logger.info(
                    f"LEVEL BUMP: call_id={call_id} "
                    f"{LEVEL_INFO_CENTRAL[current_level]['role']} -> "
                    f"{LEVEL_INFO_CENTRAL[new_level]['role']} "
                    f"(elapsed={elapsed} min, jumped {new_level - current_level} level(s))"
                )

                await update_escalation_flow(
                    call_id=0 if not is_denial_rec else call_id,  
                    alert_id=0 if is_denial_rec else str(d.get("alert_id")),
                    level=new_level,
                    event_type="LEVEL_BUMP",
                    generated_at=check_date
                )

            if cursor == 0:
                await asyncio.sleep(60)
            else:
                await asyncio.sleep(1)

        except Exception as e:
            logger.exception(f"level_bump_watcher error: {e}")
            await asyncio.sleep(10)


# ===========================================================================
# NAYA RECORD BROADCAST
# (denial_complaints_insert_worker mein insert ke baad call karo)
# ===========================================================================
async def broadcast_new_escalation(call_id: str, denial_record: dict):
    """Insert ke turant baad call karo — naya alert MDT pe broadcast hoga."""
    try:
        if await is_closed(call_id):
            return

        # 👇 FIX: Naye denial alert ka level uske naye added_date se calculate karo
        elapsed = elapsed_minutes_since(denial_record.get("added_date"))
        current_level = get_level_for_elapsed_minutes(elapsed)
        if current_level < 2:
            current_level = 2
        current_level = min(current_level, 7)
        
        # Redis me forcefully update karo
        await set_escalation_level(call_id, current_level)

        payload = await build_escalation_payload(
            call_id=call_id,
            denial_record=denial_record,
            current_level=current_level,
            event_type="new_escalation_alert",
        )
        manager.broadcast(payload)

        logger.info(
            f"NEW ESCALATION: call_id={call_id}, "
            f"level={current_level} ({LEVEL_INFO_CENTRAL[current_level]['role']})"
        )

        # =========================================================
        # 👇 TRACK IN ESCALATION FLOW TABLE (Denial ke liye alert_id = 0)
        # =========================================================
        await update_escalation_flow(
            call_id=call_id,
            alert_id=0, # 👈 Yahan 0 bhej rahe hain
            level=current_level,
            event_type="NEW_ALERT",
            generated_at=denial_record.get("added_date")
        )

    except Exception as e:
        logger.exception(f"broadcast_new_escalation FAILED: {e}")

async def broadcast_new_central_alert(central_alert_row: dict):
    """Central alert insert ke turant baad call karo —
    naya alert MDT (level 1) pe real-time broadcast hoga + FCM push hoga."""
    try:
        # 👇 FIX: Ab hum alert_id ko track karenge, incident_id ko nahi
        call_id = str(central_alert_row.get("alert_id") or "")
        if not call_id or call_id == "0":
            return

        if await is_closed(call_id):
            return

        # 👇 SLA breach details add karo before broadcast
        alerts_with_sla = await attach_sla_info_to_alerts([central_alert_row])
        if alerts_with_sla:
            central_alert_row = alerts_with_sla[0]

        # Naye alert ka level uske naye created_date se calculate karo
        elapsed = elapsed_minutes_since(central_alert_row.get("created_date"))
        current_level = get_level_for_elapsed_minutes(elapsed)
        current_level = min(current_level, 7)
        
        # Redis me forcefully update karo
        await set_escalation_level(call_id, current_level)

        payload = await build_escalation_payload(
            call_id=call_id,
            denial_record={},
            central_alerts_data=[central_alert_row],
            current_level=current_level,
            event_type="new_escalation_alert",
        )
        manager.broadcast(payload)

        logger.info(
            f"NEW CENTRAL ALERT BROADCAST: alert_id={call_id}, "
            f"ambulance={central_alert_row.get('ambulance_no')}, "
            f"level={current_level} ({LEVEL_INFO_CENTRAL[current_level]['role']})"
        )

        # 👇 NAYA FCM LOGIC: Redis se token nikal kar direct push maro
        amb_raw = central_alert_row.get("ambulance_no")
        if amb_raw:
            amb_normalized = normalize_vehicle_number(amb_raw)
            token = await redis_client.get(f"fcm_token:{amb_normalized}")
            
            if token:
                logger.info(f"🚀 Triggering FCM from worker for amb={amb_normalized}")
                await send_fcm_push(
                    token=token,
                    title=f"Ambulance {amb_raw} - New Alert",
                    body=f"Alert: {central_alert_row.get('alert_type', 'Escalation')}",
                    data={
                        "alert_id": str(call_id),
                        "type": "Late",
                    }
                )
            else:
                logger.info(f"SKIP FCM: No token found in Redis for amb={amb_normalized}")

        # =========================================================
        # 👇 TRACK IN ESCALATION FLOW TABLE (Central Alert ke liye call_id = 0)
        # =========================================================
        await update_escalation_flow(
            call_id=0,
            alert_id=str(central_alert_row.get("alert_id")),
            level=current_level,
            event_type="NEW_ALERT",
            generated_at=central_alert_row.get("created_date")
        )

    except Exception as e:
        logger.exception(f"broadcast_new_central_alert FAILED: {e}")
# ===========================================================================
# API — Take Action  (path-based level: /api/escalation/take_action/{level})
#   koi bhi role action le → incident close, aage forward nahi
# ===========================================================================
class TakeActionRequest(BaseModel):
    call_id: Optional[Union[str, int]] = None
    alert_id: Optional[Union[str, int]] = None
    action_by: Optional[str] = "unknown"
    action_by_role: Optional[str] = None
    action_remark: Optional[str] = ""
    action_type: Optional[str] = "acknowledge"

# @app.post("/api/escalation/take_action/{level}")
# async def take_escalation_action(level: int, request: Request):
#     """
#     URL:  POST /api/escalation/take_action/1    (1=MDT, 2=DM, ..., 7=CBO)

#     Body:
#     {
#         "call_id": "CALL123",
#         "action_by": "user_id",
#         "action_by_role": "MDT",
#         "action_remark": "Ambulance dispatched",
#         "action_type": "acknowledge"
#     }
#     """
#     try:
#         # Validate level
#         if level < 1 or level > 7:
#             return {
#                 "status": "error",
#                 "message": f"Invalid level: {level}. Must be 1-7.",
#                 "valid_levels": {
#                     1: "MDT", 2: "DM", 3: "ZM", 4: "OM",
#                     5: "SH", 6: "COO", 7: "CBO"
#                 }
#             }

#         action_by_level = level
#         action_by_role  = LEVEL_INFO_CENTRAL[level]["role"]

#         body              = await request.json()
#         call_id           = body.get("call_id")
#         action_by         = body.get("action_by", "unknown")
#         action_remark     = body.get("action_remark", "")
#         action_type       = body.get("action_type", "acknowledge")

#         # Body mein role diya to wahi use karo, warnha URL level se
#         if body.get("action_by_role"):
#             action_by_role = body.get("action_by_role")

#         if not call_id:
#             return {"status": "error", "message": "call_id is required"}

#         if await is_closed(call_id):
#             return {"status": "error", "message": "Incident already closed"}

#         # Mark action taken + cleanup Redis state (aage escalate nahi hoga)
#         action_details = await mark_action_taken(
#             call_id=call_id,
#             action_by=action_by,
#             action_by_role=action_by_role,
#             action_by_level=action_by_level,
#             action_remark=action_remark,
#             action_type=action_type,
#         )

#         # DB update (central_alerts → CLOSED)
#         try:
#             await database2.execute(
#                 """
#                 UPDATE public.central_alerts
#                 SET escalate_status = 'CLOSED',
#                     cancel_by        = :action_by,
#                     cancel_date      = NOW(),
#                     remark           = CONCAT(COALESCE(remark, ''),
#                                              ' | CLOSED BY ', :role, ': ', :remark)
#                 WHERE alert_id = :call_id
#                 """,
#                 {
#                     "action_by": action_by,
#                     "role":      action_by_role,
#                     "remark":    action_remark,
#                     "call_id":   call_id,
#                 },
#             )
#         except Exception as db_err:
#             logger.exception(f"DB update failed (still broadcasting): {db_err}")

#         # Broadcast to all WS clients
#         denial_record = await fetch_denial_record(call_id)
#         payload = await build_escalation_payload(
#             call_id=call_id,
#             denial_record=denial_record,
#             current_level=action_by_level,
#             event_type="escalation_closed",
#             extra={
#                 "action_by":         action_by,
#                 "action_by_role":    action_by_role,
#                 "action_by_level":   action_by_level,
#                 "action_remark":     action_remark,
#                 "action_type":       action_type,
#                 "action_taken_at":   action_details["action_taken_at"],
#                 "closed":            True,
#             },
#         )
#         manager.broadcast(payload)

#         logger.info(
#             f"ACTION TAKEN → INCIDENT CLOSED: call_id={call_id}, "
#             f"by={action_by_role} (level {action_by_level}), remark={action_remark}"
#         )

#         return {
#             "status": "success",
#             "message": "Action taken, incident closed. Aage escalate nahi hoga.",
#             "action_details": action_details,
#         }

#     except Exception as e:
#         logger.exception(f"take_escalation_action FAILED: {e}")
#         return {"status": "error", "message": str(e)}





# ============================================================
# TEST API — FCM push manually test karne ke liye
# ============================================================
@app.get("/api/test-fcm")
async def test_fcm_push(
    token: str = Query(..., description="FCM device token"),
    title: str = Query("Test Alert", description="Notification title"),
    body: str = Query("This is a test notification", description="Notification body")
):
    """
    FCM push notification manually test karne ke liye.
    
    Usage:
      GET /api/test-fcm?token=cFtOhLxNj9c:APA91b...&title=Test&body=Hello
    
    Server logs me check karo:
      ✅ FCM SUCCESS  → notification bajega
      ❌ FCM ERROR    → token ya key me problem hai
    """
    logger.info(f"TEST FCM: manual test requested for token={token[:20]}...")

    await send_fcm_push(
        token=token,
        title=title,
        body=body,
        data={"test": "true", "timestamp": str(datetime.now(ist).isoformat())}
    )

    return {
        "status": "sent",
        "message": "Check server logs for FCM response (✅ SUCCESS or ❌ ERROR)",
        "token_preview": f"{token[:20]}...",
        "title": title,
        "body": body
    }


@app.post("/api/escalation/take_action/{level}")
async def take_escalation_action(level: int, payload: TakeActionRequest):
    """
    URL:  POST /api/escalation/take_action/1    (1=MDT, 2=DM, ..., 7=CBO)

    Body:
    {
        "alert_id": 159125894,  <-- int ya string dono accept honge
        "action_by": "pilot_user_id",
        "action_remark": "Ambulance dispatched to location"
    }
    """
    try:
        # Validate level
        if level < 1 or level > 7:
            return {
                "status": "error",
                "message": f"Invalid level: {level}. Must be 1-7.",
                "valid_levels": {
                    1: "MDT", 2: "DM", 3: "ZM", 4: "OM",
                    5: "SH", 6: "COO", 7: "CBO"
                }
            }

        action_by_level = level
        action_by_role  = LEVEL_INFO_CENTRAL[level]["role"]

        # 👇 Pydantic model se data lo
        raw_id = payload.call_id if payload.call_id is not None else payload.alert_id
        
        # ID ko hamesha string me convert karo (agar int aaya to str ban jayega)
        payload_id = str(raw_id).strip() if raw_id is not None else ""
        
        action_by         = payload.action_by if payload.action_by else "unknown"
        action_remark     = payload.action_remark if payload.action_remark else ""
        action_type       = payload.action_type if payload.action_type else "acknowledge"

        if payload.action_by_role:
            action_by_role = payload.action_by_role

        # 👇 DEBUG LOG
        logger.info(f"🚀 TAKE ACTION API CALLED. Received ID: '{payload_id}'")

        if not payload_id or payload_id.lower() == "none":
            return {"status": "error", "message": "call_id or alert_id is required"}

        # Check if already closed
        is_already_closed = await is_closed(payload_id)
        logger.info(f"🔍 Checking is_closed for ID '{payload_id}': {is_already_closed}")
        
        if is_already_closed:
            logger.warning(f"❌ BLOCKED: Incident {payload_id} is already closed in Redis.")
            return {"status": "error", "message": "Incident already closed"}

        # Mark action taken + cleanup Redis state (aage escalate nahi hoga)
        action_details = await mark_action_taken(
            call_id=payload_id,
            action_by=action_by,
            action_by_role=action_by_role,
            action_by_level=action_by_level,
            action_remark=action_remark,
            action_type=action_type,
        )

        # =========================================================
        # IDENTIFY & UPDATE CORRECT TABLE
        # =========================================================
        # 👇 ID ko int me convert karke query me daal rahe hain taaki DB match ho jaye
        try:
            db_id = int(payload_id)
        except ValueError:
            db_id = payload_id

        central_record = await database2.fetch_one(
            "SELECT alert_id FROM public.central_alerts WHERE alert_id = :id",
            {"id": db_id}
        )

        if central_record:
            # --- CENTRAL ALERT ---
            try:
                await database2.execute(
                    """
                    UPDATE public.central_alerts
                    SET escalate_status = '2',
                        escalated_deny_remark = :remark,
                        cancel_by = :action_by,
                        cancel_date = NOW(),
                        updated_date = NOW()
                    WHERE alert_id = :alert_id
                    """,
                    {
                        "remark": action_remark,
                        "action_by": action_by,
                        "alert_id": db_id,
                    },
                )
                logger.info(f"✅ CENTRAL TABLE UPDATED: alert_id={payload_id}, status=2")
            except Exception as db_err:
                logger.exception(f"Central DB update failed: {db_err}")
        else:
            # --- DENIAL ALERT ---
            try:
                await database2.execute(
                    """
                    UPDATE public.denial_escalation_master
                    SET escalate_status = '2',
                        remark = :remark
                    WHERE call_id = :call_id
                    """,
                    {
                        "remark": action_remark,
                        "call_id": str(db_id), # Denial table me call_id varchar hai
                    },
                )
                logger.info(f"✅ DENIAL TABLE UPDATED: call_id={payload_id}, status=2")
            except Exception as db_err:
                logger.exception(f"Denial DB update failed: {db_err}")

        # =========================================================
        # BROADCAST TO ALL WS CLIENTS
        # =========================================================
        denial_record_payload = await fetch_denial_record(payload_id)
        broadcast_payload = await build_escalation_payload(
            call_id=payload_id,
            denial_record=denial_record_payload,
            current_level=action_by_level,
            event_type="escalation_closed",
            extra={
                "action_by":         action_by,
                "action_by_role":    action_by_role,
                "action_by_level":   action_by_level,
                "action_remark":     action_remark,
                "action_type":       action_type,
                "action_taken_at":   action_details["action_taken_at"],
                "closed":            True,
            },
        )
        manager.broadcast(broadcast_payload)

        logger.info(
            f"ACTION TAKEN → INCIDENT CLOSED: id={payload_id}, "
            f"by={action_by_role} (level {action_by_level}), remark={action_remark}"
        )

        # =========================================================
        # TRACK IN AUDIT TABLE (alert_escalation_flow)
        # =========================================================
        await update_escalation_flow(
            call_id=payload_id,
            alert_id=payload_id, 
            level=action_by_level,
            event_type="ACTION_TAKEN",
            action_by=action_by,
            action_remark=action_remark
        )

        return {
            "status": "success",
            "message": "Action taken, incident closed successfully.",
            "action_details": action_details,
        }

    except Exception as e:
        logger.exception(f"take_escalation_action FAILED: {e}")
        return {"status": "error", "message": str(e)}




@app.get("/api/mdt/alerts")
async def get_mdt_alerts_by_ambulance(ambulance_no: str = Query(..., description="Ambulance Number")):
    """
    Fetch all alerts for a specific ambulance (MDT Level 1 perspective).
    
    URL: GET /api/mdt/alerts?ambulance_no=TT-00-MP-0001
    
    Returns:
    - Direct data array (no WebSocket INITIAL_LOAD wrapper).
    - Includes all past and present alerts.
    - Adds 'able_to_action' boolean (True if < 10 min old & not closed).
    """
    try:
        normalized_amb = normalize_vehicle_number(ambulance_no)
        
        # SQL query to fetch all alerts for this ambulance (normalized match)
        query = """
            SELECT * FROM public.central_alerts
            WHERE UPPER(
                REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(ambulance_no, ' ', ''), '-', ''), '.', ''), '_', ''), '/', '')
            ) = :amb
            ORDER BY created_date DESC
            LIMIT 500
        """
        
        rows = await database2.fetch_all(query, {"amb": normalized_amb})
        
        if not rows:
            # 👇 Direct format
            return {
                "data": [],
                "vehicle_filter": normalized_amb
            }

        # Serialize rows
        alerts = [serialize_row(r) for r in rows]
        # Attach SLA info to all alerts
        alerts = await attach_sla_info_to_alerts(alerts)
        
        response_data = []
        for alert in alerts:
            call_id = str(alert.get("alert_id") or "")
            check_date = alert.get("created_date")
            
            # Check if closed
            if await is_closed(call_id):
                current_level = 1  # Default fallback
                is_closed_flag = True
            else:
                # Calculate current level based on time
                elapsed = elapsed_minutes_since(check_date)
                current_level = get_level_for_elapsed_minutes(elapsed)
                is_closed_flag = False
                
            alert["current_level"] = current_level
            alert["current_role"] = LEVEL_INFO_CENTRAL.get(current_level, {}).get("role", "MDT")
            alert["is_closed"] = is_closed_flag
            
            # 👇 MAIN LOGIC: able_to_action
            # True only if alert is Level 1 (under 10 min) AND not closed
            alert["able_to_action"] = (current_level == 1 and not is_closed_flag)
            
            response_data.append(alert)
            
        # 👇 Direct format (No type, No count)
        return {
            "data": response_data,
            "vehicle_filter": normalized_amb
        }

    except Exception as e:
        logger.exception(f"get_mdt_alerts_by_ambulance FAILED: {e}")
        return {"status": "error", "message": str(e)}